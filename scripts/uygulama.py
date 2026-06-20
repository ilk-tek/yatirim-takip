# ==========================================
# YATIRIM TAKİP — ANA UYGULAMA
# ==========================================

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
os.chdir(os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import sqlite3
import pandas as pd
from db import baglan, senkronize_et, sql_oku
from datetime import date, timedelta
from hesaplamalar import (
    performans_ozeti, twr_hesapla, yilliklandir,
    mevduat_deger_hesapla, aylik_portfoy_ozeti,
    aylik_dagilim_hesapla, AY_ISIMLERI,
    fifo_maliyet_hesapla, MEVDUAT_TURLERI,
    kur_getir, bugunun_kuru,
    donemsel_karsilastirma_hesapla
)
from fiyat_cek import hisse_fiyatlari_cek, bist_fiyatlari_cek, kripto_fiyatlari_cek, altin_fiyatlari_cek, tum_fiyatlari_cek
from tefas_import import tefas_import
from kur_guncelle import kur_cek_ve_kaydet

# --- Veritabanı bağlantısı ---
def veritabani_baglan():
    return baglan()

# ==========================================
# YARDIMCI: Aracı Kurum ve Portföy Etiketi Listeleri
# ==========================================
# Tablolar yoksa (eski kurulum) boş liste döner,
# işlem Ekle/Düzenle sayfaları sabit listeden devam eder.

def araci_kurum_listesi():
    """Veritabanındaki aracı kurumları [''] + sıralı liste olarak döndürür."""
    try:
        import db
        db._baglanti = None          # taze bağlantı garantile
        df = sql_oku("SELECT ad FROM araci_kurumlar ORDER BY ad", baglan())
        return [""] + df["ad"].tolist()
    except Exception:
        return ["", "İş Yatırım", "İş Bankası", "YKB", "Anadolubank",
                "Ata Yatırım", "Garanti BBVA", "Akbank", "Kiralık Kasa", "Midas"]

def portfoy_etiketi_listesi():
    """Veritabanındaki portföy etiketlerini [''] + sıralı liste olarak döndürür."""
    try:
        import db
        db._baglanti = None
        df = sql_oku("SELECT ad FROM portfoy_etiketleri ORDER BY ad", baglan())
        return [""] + df["ad"].tolist()
    except Exception:
        return ["", "Yatırım", "Defans", "Atak", "YP Fon",
                "Arbitraj", "Emtia", "Uzun Borçlanma", "M"]

def araci_kurum_kaydet(ad):
    """Yeni aracı kurumu veritabanına ekler."""
    import db
    db._baglanti = None          # eski bağlantıyı bırak
    baglanti = baglan()           # taze bağlantı
    cursor   = baglanti.cursor()
    # Tablo yoksa oluştur
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS araci_kurumlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ad TEXT NOT NULL UNIQUE
        )
    """)
    cursor.execute("INSERT OR IGNORE INTO araci_kurumlar (ad) VALUES (?)", (ad.strip(),))
    baglanti.commit()
    senkronize_et()
    db._baglanti = None          # tekrar sıfırla (sonraki okumaları korur)

def portfoy_etiketi_kaydet(ad):
    """Yeni portföy etiketini veritabanına ekler."""
    import db
    db._baglanti = None
    baglanti = baglan()
    cursor   = baglanti.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS portfoy_etiketleri (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ad TEXT NOT NULL UNIQUE
        )
    """)
    cursor.execute("INSERT OR IGNORE INTO portfoy_etiketleri (ad) VALUES (?)", (ad.strip(),))
    baglanti.commit()
    senkronize_et()
    db._baglanti = None

# --- Sayfa ayarları ---
st.set_page_config(
    page_title="Yatırım Takip",
    page_icon="📈",
    layout="wide"
)

# ==========================================
# ŞİFRE KORUMASI
# ==========================================
# admin_sifre  → tam erişim (sen)
# izleyici_sifre → sadece görüntüleme (diğer kişi)
# Şifreler Streamlit Cloud'da "secrets" bölümüne girilecek.

ADMIN_SIFRE    = st.secrets.get("ADMIN_SIFRE", "")
IZLEYICI_SIFRE = st.secrets.get("IZLEYICI_SIFRE", "")

if "erisim_seviyesi" not in st.session_state:
    st.session_state.erisim_seviyesi = None  # henüz giriş yapılmadı

if st.session_state.erisim_seviyesi is None:
    st.title("🔐 Yatırım Takip")
    girilen = st.text_input("Şifre:", type="password")
    if st.button("Giriş"):
        if girilen == ADMIN_SIFRE:
            st.session_state.erisim_seviyesi = "admin"
            st.rerun()
        elif girilen == IZLEYICI_SIFRE:
            st.session_state.erisim_seviyesi = "izleyici"
            st.rerun()
        else:
            st.error("Yanlış şifre!")
    st.stop()  # şifre girilmeden aşağısı çalışmaz

# Kısayol: hangi modda olduğumuzu bir değişkene atalım
ADMIN_MOD = (st.session_state.erisim_seviyesi == "admin")

# --- Kenar çubuğu menü ---
# Admin modunda tüm sayfalar görünür
# İzleyici modunda sadece görüntüleme sayfaları görünür

if ADMIN_MOD:
    menu_secenekleri = [
        "📊 Portföy",
        "🏦 Aracı Kurum",
        "📈 Performans",
        "📅 Aylık Özet",
        "🔄 Dönemsel Karşılaştırma",
        "🏛️ Yatırım Fonları",
        "💱 Fiyat Güncelle",
        "➕ Varlık Ekle",
        "✏️ Varlık Düzenle",
        "💰 İşlem Ekle",
        "✏️ İşlem Düzenle",
        "📤 İşlem Yükle",
        "📋 İşlem Geçmişi",
        "🗓️ Fiyat Geçmişi",
    ]
else:
    menu_secenekleri = [
        "📊 Portföy",
        "🏦 Aracı Kurum",
        "📈 Performans",
        "📅 Aylık Özet",
        "🔄 Dönemsel Karşılaştırma",
        "🏛️ Yatırım Fonları",
        "📋 İşlem Geçmişi",
    ]

sayfa = st.sidebar.radio("Menü", menu_secenekleri)

# --- Bulut senkronizasyon butonu (sadece admin görür) ---
if ADMIN_MOD:
    st.sidebar.markdown("---")
    if st.sidebar.button("🔄 Bulut ile Senkronize Et"):
        if senkronize_et():
            st.sidebar.success("Senkronize edildi!")
        else:
            st.sidebar.warning("Senkronizasyon yapılamadı (bağlantıyı kontrol edin).")

# ==========================================
# SAYFA 1: PORTFÖY
# ==========================================
if sayfa == "📊 Portföy":
    st.title("📊 Portföyüm")
    st.markdown("---")

    baglanti = veritabani_baglan()

    df = sql_oku("""
        SELECT
            v.id,
            v.kod,
            v.ad,
            v.tur,
            v.para_birimi,
            v.exposure,
            SUM(CASE WHEN i.islem_turu = 'Alış'  THEN i.adet
                         WHEN i.islem_turu = 'Satış' THEN -i.adet
                         ELSE 0 END) AS toplam_adet,
            SUM(CASE WHEN i.islem_turu = 'Alış' THEN i.adet  ELSE 0 END)        AS toplam_alis_adet,
            SUM(CASE WHEN i.islem_turu = 'Alış' THEN i.tutar ELSE 0 END)        AS toplam_alis_tutar
        FROM varliklar v
        LEFT JOIN islemler i ON v.id = i.varlik_id
        GROUP BY v.id
    """, baglanti)
    son_fiyatlar = sql_oku("""
        SELECT f1.varlik_id, f1.fiyat
        FROM fiyat_gecmisi f1
        INNER JOIN (
            SELECT varlik_id, MAX(tarih) AS son_tarih
            FROM fiyat_gecmisi
            GROUP BY varlik_id
        ) f2 ON f1.varlik_id = f2.varlik_id AND f1.tarih = f2.son_tarih
    """, baglanti)

    if df.empty:
        st.info("Henüz varlık eklenmemiş.")
    else:
        # Son fiyat güncelleme tarihini göster
        son_guncelleme = sql_oku("""
            SELECT MAX(tarih) as tarih FROM fiyat_gecmisi
        """, veritabani_baglan()).iloc[0]["tarih"]
        if son_guncelleme:
            st.info(f"📅 Son fiyat güncellemesi: **{son_guncelleme}** — Güncellemek için 💱 Fiyat Güncelle sayfasına gidin.")
        else:
            st.warning("Henüz fiyat girilmemiş. 💱 Fiyat Güncelle sayfasından fiyat girin.")

        # --- Fiyat geçmişi eksik varlık uyarısı ---
        eksik_fiyat = sql_oku("""
            SELECT
                v.kod, v.tur,
                MIN(i.tarih) AS ilk_islem,
                (SELECT MIN(f.tarih) FROM fiyat_gecmisi f WHERE f.varlik_id = v.id) AS ilk_fiyat
            FROM varliklar v
            JOIN islemler i ON v.id = i.varlik_id AND i.islem_turu = 'Alış'
            JOIN (
                SELECT varlik_id, SUM(CASE WHEN islem_turu = 'Alış' THEN adet ELSE -adet END) AS net
                FROM islemler WHERE islem_turu IN ('Alış', 'Satış')
                GROUP BY varlik_id HAVING net > 0
            ) p ON v.id = p.varlik_id
            WHERE v.tur NOT IN ('TL Mevduat', 'YP Mevduat')
            GROUP BY v.id
        """, veritabani_baglan())

        if not eksik_fiyat.empty:
            uyarilar = []
            for _, r in eksik_fiyat.iterrows():
                if r["ilk_fiyat"] is None or pd.isna(r["ilk_fiyat"]):
                    uyarilar.append(f"**{r['kod']}** ({r['tur']}) — ilk işlem: {r['ilk_islem']}, fiyat verisi yok")
                else:
                    from datetime import datetime as _dt
                    islem_dt = _dt.strptime(r["ilk_islem"], "%Y-%m-%d")
                    fiyat_dt = _dt.strptime(r["ilk_fiyat"], "%Y-%m-%d")
                    bosluk = (fiyat_dt - islem_dt).days
                    if bosluk > 3:
                        uyarilar.append(f"**{r['kod']}** ({r['tur']}) — ilk işlem: {r['ilk_islem']}, ilk fiyat: {r['ilk_fiyat']} ({bosluk} gün boşluk)")
            if uyarilar:
                st.warning("⚠️ Fiyat verisi eksik:\n\n" + "\n\n".join(uyarilar) + "\n\n💱 Fiyat Güncelle → Geçmiş Veri Tamamla bölümünden çekin.")

        st.markdown("---")

        # ==========================================
        # BUGÜNKÜ DÖVİZ KURLARI (bir kere çek, hep kullan)
        # ==========================================
        usd_kuru_bugun = bugunun_kuru("USD")
        eur_kuru_bugun = bugunun_kuru("EUR")
        gbp_kuru_bugun = bugunun_kuru("GBP")

        # Kurları kenar çubuğunda göster
        st.sidebar.markdown("---")
        st.sidebar.markdown("**💱 Güncel Kurlar**")
        st.sidebar.markdown(f"USD/TRY: **{usd_kuru_bugun:.4f}**")
        st.sidebar.markdown(f"EUR/TRY: **{eur_kuru_bugun:.4f}**")
        st.sidebar.markdown(f"GBP/TRY: **{gbp_kuru_bugun:.4f}**")

        # Fiyatları son kayıtlardan al
        guncel_fiyatlar = {}
        for _, row in son_fiyatlar.iterrows():
            guncel_fiyatlar[row["varlik_id"]] = row["fiyat"]

        ozet                   = []
        portfoy_toplam_maliyet = 0
        toplam_deger           = 0

        for _, row in df.iterrows():
            guncel_fiyat = guncel_fiyatlar.get(row["id"], 0)
            para_birimi  = row["para_birimi"] if row["para_birimi"] else "TRY"
            kur          = bugunun_kuru(para_birimi)
            is_mevduat   = row["tur"] in MEVDUAT_TURLERI

            # Net adet kontrolü
            if not (row["toplam_adet"] and row["toplam_adet"] > 0):
                continue

            if is_mevduat:
                # ==========================================
                # MEVDUAT: fiyat her zaman 1
                # adet = bakiye tutarı, değer = adet × kur
                # fiyat_gecmisi'ne gerek yok
                # ==========================================
                guncel_deger     = row["toplam_adet"] * 1.0 * kur
                pozisyon_maliyet = fifo_maliyet_hesapla(row["id"])
            else:
                # ==========================================
                # NORMAL VARLIK: fiyat_gecmisi'nden gelir
                # ==========================================
                if guncel_fiyat <= 0:
                    continue
                guncel_deger     = row["toplam_adet"] * guncel_fiyat * kur
                pozisyon_maliyet = fifo_maliyet_hesapla(row["id"])

            # ==========================================
            # KÂR/ZARAR VE ÖZET
            # ==========================================
            kar_zarar    = guncel_deger - pozisyon_maliyet
            yuzde_getiri = (kar_zarar / pozisyon_maliyet * 100) if pozisyon_maliyet else 0

            portfoy_toplam_maliyet += pozisyon_maliyet
            toplam_deger           += guncel_deger

            ozet.append({
                "Kod"            : row["kod"],
                "Ad"             : row["ad"],
                "Tür"            : row["tur"],
                "PB"             : para_birimi,
                "Exposure"       : row["exposure"],
                "Adet"           : row["toplam_adet"],
                "Maliyet (TL)"   : pozisyon_maliyet,
                "Değer (TL)"     : guncel_deger,
                "Kâr/Zarar (TL)" : kar_zarar,
                "Getiri %"       : yuzde_getiri,
                "Değer (USD)"    : guncel_deger / usd_kuru_bugun,
            })

        if ozet:
            ozet_df = pd.DataFrame(ozet)

            # ==========================================
            # TOPLAM METRİKLER (en üstte)
            # ==========================================
            toplam_kar    = toplam_deger - portfoy_toplam_maliyet
            toplam_getiri = (toplam_kar / portfoy_toplam_maliyet * 100) if portfoy_toplam_maliyet else 0

            col1, col2, col3 = st.columns(3)
            col1.metric("💼 Toplam Maliyet",     f"{portfoy_toplam_maliyet:,.0f} TL")
            col2.metric("📈 Toplam Güncel Değer", f"{toplam_deger:,.0f} TL")
            col3.metric("💰 Toplam Kâr/Zarar",    f"{toplam_kar:,.0f} TL",
                        delta=f"{toplam_getiri:.2f}%")

            col4, col5, col6 = st.columns(3)
            col4.metric("💵 Maliyet (USD)",       f"${portfoy_toplam_maliyet / usd_kuru_bugun:,.0f}")
            col5.metric("💵 Güncel Değer (USD)",   f"${toplam_deger / usd_kuru_bugun:,.0f}")
            col6.metric("💵 Kâr/Zarar (USD)",     f"${toplam_kar / usd_kuru_bugun:,.0f}",
                        delta=f"{toplam_getiri:.2f}%")

            st.caption("ℹ️ Gerçekleşmemiş (unrealized) kâr/zarar. Maliyet: FIFO yöntemi, alış tarihindeki döviz kuru ile TL'ye çevrilir. Güncel Değer: adet × son fiyat × bugünkü kur.")

            # ==========================================
            # HİYERARŞİK PORTFÖY ÖZETİ: Exposure → Tür → Varlık
            # ==========================================
            st.markdown("---")
            st.subheader("Portföy Özeti")

            # Varlık bazlı aracı kurum bilgisi (en çok kullanılan)
            baglanti_ak = veritabani_baglan()
            ak_bilgi = sql_oku("""
                SELECT varlik_id,
                       COALESCE(araci_kurum, '') AS araci_kurum,
                       SUM(CASE WHEN islem_turu = 'Alış' THEN adet ELSE -adet END) AS net
                FROM islemler
                WHERE islem_turu IN ('Alış', 'Satış')
                GROUP BY varlik_id, araci_kurum
                HAVING net > 0
            """, baglanti_ak)

            # Varlık → aracı kurum eşlemesi (en büyük pozisyon)
            varlik_araci = {}
            if not ak_bilgi.empty:
                for vid in ak_bilgi["varlik_id"].unique():
                    vid_df = ak_bilgi[ak_bilgi["varlik_id"] == vid]
                    en_buyuk = vid_df.loc[vid_df["net"].idxmax()]
                    ak = en_buyuk["araci_kurum"]
                    varlik_araci[vid] = ak if ak else "Belirtilmemiş"

            # ozet_df'e aracı kurum ekle
            ozet_df["Aracı Kurum"] = ozet_df.apply(
                lambda r: varlik_araci.get(
                    df[df["kod"] == r["Kod"]]["id"].values[0], "Belirtilmemiş"
                ) if len(df[df["kod"] == r["Kod"]]) > 0 else "Belirtilmemiş",
                axis=1
            )

            portfoy_toplam_tl = ozet_df["Değer (TL)"].sum()

            # --- Yardımcı fonksiyonlar ---
            def pf_deger_str(filtre_df, toplam):
                """Değer + pay string'i oluşturur."""
                tl  = filtre_df["Değer (TL)"].sum()
                usd = filtre_df["Değer (USD)"].sum()
                pay = (tl / toplam * 100) if toplam else 0
                return (
                    f"**{tl:,.0f} TL** | "
                    f"**${usd:,.0f}** | "
                    f"Pay: **%{pay:.1f}**"
                )

            # --- Exposure → Tür → Varlık expander'ları ---
            exposurelar = sorted(ozet_df["Exposure"].unique())

            for exp in exposurelar:
                exp_df = ozet_df[ozet_df["Exposure"] == exp]

                with st.expander(f"📂 {exp}  —  {pf_deger_str(exp_df, portfoy_toplam_tl)}"):
                    turler = sorted(exp_df["Tür"].unique())

                    for tur in turler:
                        tur_df = exp_df[exp_df["Tür"] == tur]

                        with st.expander(f"📄 {tur}  —  {pf_deger_str(tur_df, portfoy_toplam_tl)}"):
                            goster = tur_df[["Kod", "Ad", "PB", "Exposure",
                                             "Adet", "Değer (TL)", "Değer (USD)",
                                             "Aracı Kurum"]].copy()
                            goster = goster.sort_values("Değer (TL)", ascending=False)
                            st.dataframe(
                                goster.style.format({
                                    "Adet"        : "{:,.4f}",
                                    "Değer (TL)"  : "{:,.0f}",
                                    "Değer (USD)" : "${:,.0f}",
                                }),
                                use_container_width=True, hide_index=True
                            )

        else:
            st.info("Henüz işlem girilmemiş.")

# ==========================================
# SAYFA: ARACI KURUM BAZINDA PORTFÖY
# ==========================================
elif sayfa == "🏦 Aracı Kurum":
    st.title("🏦 Aracı Kurum Bazında Portföy")
    st.markdown("---")

    baglanti = veritabani_baglan()

    # ==========================================
    # BUGÜNKÜ DÖVİZ KURLARI
    # ==========================================
    usd_kuru_bugun = bugunun_kuru("USD")

    # Kurları kenar çubuğunda göster
    eur_kuru_bugun = bugunun_kuru("EUR")
    gbp_kuru_bugun = bugunun_kuru("GBP")
    st.sidebar.markdown("---")
    st.sidebar.markdown("**💱 Güncel Kurlar**")
    st.sidebar.markdown(f"USD/TRY: **{usd_kuru_bugun:.4f}**")
    st.sidebar.markdown(f"EUR/TRY: **{eur_kuru_bugun:.4f}**")
    st.sidebar.markdown(f"GBP/TRY: **{gbp_kuru_bugun:.4f}**")

    # Son fiyatları çek
    son_fiyatlar_ak = sql_oku("""
        SELECT f1.varlik_id, f1.fiyat
        FROM fiyat_gecmisi f1
        INNER JOIN (
            SELECT varlik_id, MAX(tarih) AS son_tarih
            FROM fiyat_gecmisi
            GROUP BY varlik_id
        ) f2 ON f1.varlik_id = f2.varlik_id AND f1.tarih = f2.son_tarih
    """, baglanti)

    guncel_fiyatlar_ak = {}
    for _, row in son_fiyatlar_ak.iterrows():
        guncel_fiyatlar_ak[row["varlik_id"]] = row["fiyat"]

    # İşlem bazlı veri: aynı varlığın farklı aracı kurumlardaki
    # pozisyonlarını AYRI gösterir.
    # Sadece Değer (TL) ve Değer (USD) — Maliyet/K-Z yok
    # (çünkü FIFO aracı kurum bazında ayrıştırılamaz).
    islem_bazli = sql_oku("""
        SELECT
            v.id   AS varlik_id,
            v.kod, v.ad, v.tur, v.para_birimi, v.exposure,
            COALESCE(i.araci_kurum, '')      AS araci_kurum_ham,
            SUM(CASE WHEN i.islem_turu = 'Alış' THEN i.adet
                     ELSE -i.adet END)       AS net_adet
        FROM islemler i
        JOIN varliklar v ON v.id = i.varlik_id
        WHERE i.islem_turu IN ('Alış', 'Satış')
        GROUP BY v.id, i.araci_kurum
        HAVING net_adet > 0
    """, baglanti)

    if islem_bazli.empty:
        st.info("Henüz işlem girilmemiş.")
    else:
        # Boş aracı kurum → "Belirtilmemiş"
        islem_bazli["Aracı Kurum"] = islem_bazli["araci_kurum_ham"].apply(
            lambda x: x if x else "Belirtilmemiş"
        )

        # Her satır için değer hesapla
        deger_tl  = []
        deger_usd = []
        for _, r in islem_bazli.iterrows():
            pb  = r["para_birimi"] if r["para_birimi"] else "TRY"
            kur = bugunun_kuru(pb)

            if r["tur"] in MEVDUAT_TURLERI:
                fiyat = 1.0
            else:
                fiyat = guncel_fiyatlar_ak.get(r["varlik_id"], 0)

            d = r["net_adet"] * fiyat * kur
            deger_tl.append(d)
            deger_usd.append(d / usd_kuru_bugun)

        islem_bazli["Değer (TL)"]  = deger_tl
        islem_bazli["Değer (USD)"] = deger_usd

        # --- Yardımcı fonksiyonlar ---
        def deger_ozet(df_filtre):
            """Değer toplamlarını döndürür."""
            return {
                "Değer (TL)"  : df_filtre["Değer (TL)"].sum(),
                "Değer (USD)" : df_filtre["Değer (USD)"].sum(),
            }

        def deger_str(oz):
            """Özet sözlüğünü okunabilir string'e çevirir."""
            return (
                f"Değer: **{oz['Değer (TL)']:,.0f} TL** | "
                f"USD: **${oz['Değer (USD)']:,.0f}**"
            )

        # --- Expander'lar ---
        araci_kurumlar = sorted(islem_bazli["Aracı Kurum"].unique())

        for araci in araci_kurumlar:
            ak_df = islem_bazli[islem_bazli["Aracı Kurum"] == araci]
            ak_oz = deger_ozet(ak_df)

            with st.expander(f"🏦 {araci}  —  {deger_str(ak_oz)}"):
                turler = sorted(ak_df["tur"].unique())

                for tur in turler:
                    tur_df = ak_df[ak_df["tur"] == tur]
                    tur_oz = deger_ozet(tur_df)

                    with st.expander(f"📂 {tur}  —  {deger_str(tur_oz)}"):
                        # Exposure özet tablosu
                        exp_grp = tur_df.groupby("exposure").agg(
                            DegerTL=("Değer (TL)", "sum"),
                            DegerUSD=("Değer (USD)", "sum"),
                        ).reset_index()
                        exp_grp.columns = ["Exposure", "Değer (TL)", "Değer (USD)"]

                        st.markdown("**Exposure Özeti**")
                        st.dataframe(
                            exp_grp.style.format({
                                "Değer (TL)"  : "{:,.0f}",
                                "Değer (USD)" : "${:,.0f}",
                            }),
                            use_container_width=True
                        )

                        # Varlık listesi
                        st.markdown("**Varlık Listesi**")
                        goster = tur_df[["kod", "ad", "para_birimi", "exposure",
                                          "net_adet", "Değer (TL)", "Değer (USD)"]].copy()
                        goster.columns = ["Kod", "Ad", "PB", "Exposure",
                                          "Adet", "Değer (TL)", "Değer (USD)"]
                        st.dataframe(
                            goster.style.format({
                                "Adet"        : "{:,.4f}",
                                "Değer (TL)"  : "{:,.0f}",
                                "Değer (USD)" : "${:,.0f}",
                            }),
                            use_container_width=True
                        )


# ==========================================
# SAYFA 2: PERFORMANS
# ==========================================
elif sayfa == "📈 Performans":
    st.title("📈 Performans Raporu")
    st.markdown("---")

    # ==========================================
    # TWR AÇIKLAMASI (detaylı)
    # ==========================================
    with st.expander("ℹ️ TWR (Time-Weighted Return) Nedir?", expanded=False):
        st.markdown("""
**TWR (Time-Weighted Return — Zaman Ağırlıklı Getiri)**, bir varlığın fiyat performansını ölçer.
Yapılan yatırım tutarından bağımsız olarak varlıkların performanslarını birbiriyle karşılaştırmak
amacıyla kullanılır. Örneğin 10.000 TL yatırım yapılan bir fon ile 100.000 TL yatırım yapılan
bir hisse senedinin getiri oranları, yatırım büyüklüğünden etkilenmeden karşılaştırılabilir.

**Nasıl hesaplanır?**
- Varlığın `fiyat_gecmisi` tablosundaki tüm fiyat kayıtları kronolojik sıralanır.
- Ardışık iki fiyat arasındaki oran hesaplanır (günlük getiri = bugünkü fiyat / dünkü fiyat).
- Tüm günlük getiriler birbiriyle çarpılarak toplam getiri bulunur (chain-linking yöntemi).

**TWR hangi tarihten itibaren hesaplanır?**
- TWR, varlığın **ilk fiyat kaydından son fiyat kaydına** kadar hesaplanır.
- Bir varlığı alıp satıp tekrar aldığınızda, fiyat geçmişi kesintisiz devam ettiği sürece
  TWR tüm dönem boyunca hesaplanır. Pozisyon kapanması TWR hesabını sıfırlamaz.

**Fiyat bilgisi olmayan günler ne olur?**
- Sadece `fiyat_gecmisi` tablosunda kaydı olan günler hesaplamaya dahil edilir.
- Fiyat kaydı olmayan günler atlanır; bu durum TWR sonucunu etkilemez çünkü
  mevcut iki fiyat arasındaki oran zaten o aradaki toplam değişimi yansıtır.

**TWR (TL) ve TWR (PB) farkı:**
- **TWR (TL):** Fiyat her gün o günkü döviz kuru ile TL'ye çevrilir, kur etkisi dahildir.
- **TWR (PB):** Varlığın kendi para birimindeki (USD, EUR vb.) fiyat değişimi. TRY varlıklarda ikisi aynıdır.

**Mevduat hesapları için TWR hesaplanmaz.**
""")

    donem = st.selectbox("Dönem seçin:", [
        "bu_ay", "son_3_ay", "son_6_ay", "bu_yil", "tum_zamanlar"
    ], format_func=lambda x: {
        "bu_ay"        : "Bu Ay",
        "son_3_ay"     : "Son 3 Ay",
        "son_6_ay"     : "Son 6 Ay",
        "bu_yil"       : "Bu Yıl",
        "tum_zamanlar" : "Tüm Zamanlar"
    }[x])

    st.markdown("---")

    performans_df = performans_ozeti(donem)

    if performans_df.empty:
        st.info("Seçilen dönem için yeterli fiyat verisi yok.")
    else:
        st.subheader("Tür Bazında Özet")

        performans_df["TWR_TL_sayi"] = performans_df["TWR % (TL)"].str.replace("%", "").astype(float)
        performans_df["TWR_PB_sayi"] = performans_df["TWR % (PB)"].str.replace("%", "").astype(float)

        tur_ozet = performans_df.groupby("Tür").agg(
            Varlık_Sayısı=("Kod", "count"),
            Ort_TWR_TL=("TWR_TL_sayi", "mean"),
            Ort_TWR_PB=("TWR_PB_sayi", "mean")
        ).reset_index()
        tur_ozet["Ort TWR % (TL)"] = tur_ozet["Ort_TWR_TL"].apply(lambda x: f"{x:.2f}%")
        tur_ozet["Ort TWR % (PB)"] = tur_ozet["Ort_TWR_PB"].apply(lambda x: f"{x:.2f}%")

        st.dataframe(tur_ozet[["Tür", "Varlık_Sayısı", "Ort TWR % (TL)", "Ort TWR % (PB)"]], use_container_width=True)

        st.markdown("---")
        st.subheader("Tür Bazında Detay")

        for tur in performans_df["Tür"].unique():
            tur_df   = performans_df[performans_df["Tür"] == tur]
            eski     = len(tur_df[tur_df["Güncelleme"] > 7])
            baslik   = f"📂 {tur} — {len(tur_df)} varlık"
            if eski > 0:
                baslik += f"  ⚠️ {eski} varlığın fiyatı 7+ gün eski"
            with st.expander(baslik):
                def renk_tur(row):
                    if row["Güncelleme"] > 30:
                        return ["background-color: #FFE0E0"] * len(row)
                    elif row["Güncelleme"] > 7:
                        return ["background-color: #FFF9C4"] * len(row)
                    return [""] * len(row)
                goster = tur_df[["Kod", "Ad", "PB", "TWR % (TL)", "TWR % (PB)", "Yıllık (TL)", "Yıllık (PB)", "Son Fiyat", "Güncelleme"]].copy()
                st.dataframe(goster.style.apply(renk_tur, axis=1), use_container_width=True)

        st.markdown("---")
        toplam_twr_tl = performans_df["TWR_TL_sayi"].mean()
        toplam_twr_pb = performans_df["TWR_PB_sayi"].mean()
        col1, col2 = st.columns(2)
        col1.metric("📊 Portföy Ort. TWR (TL)", f"{toplam_twr_tl:.2f}%")
        col2.metric("📊 Portföy Ort. TWR (PB)", f"{toplam_twr_pb:.2f}%")

        st.markdown("---")
        st.subheader("Exposure Bazında Özet")

        baglanti      = veritabani_baglan()
        exposure_bilgi = sql_oku("SELECT kod AS Kod, exposure FROM varliklar", baglanti)

        performans_exp = performans_df.merge(exposure_bilgi, on="Kod", how="left")

        exp_ozet = performans_exp.groupby("exposure").agg(
            Varlık_Sayısı=("Kod", "count"),
            Ort_TWR_TL=("TWR_TL_sayi", "mean"),
            Ort_TWR_PB=("TWR_PB_sayi", "mean")
        ).reset_index()
        exp_ozet.columns = ["Exposure", "Varlık Sayısı", "Ort_TWR_TL", "Ort_TWR_PB"]
        exp_ozet["Ort TWR % (TL)"] = exp_ozet["Ort_TWR_TL"].apply(lambda x: f"{x:.2f}%")
        exp_ozet["Ort TWR % (PB)"] = exp_ozet["Ort_TWR_PB"].apply(lambda x: f"{x:.2f}%")

        st.dataframe(exp_ozet[["Exposure", "Varlık Sayısı", "Ort TWR % (TL)", "Ort TWR % (PB)"]], use_container_width=True)

        st.markdown("---")
        st.subheader("Exposure Bazında Detay")

        for exp in performans_exp["exposure"].unique():
            exp_df = performans_exp[performans_exp["exposure"] == exp]
            eski   = len(exp_df[exp_df["Güncelleme"] > 7])
            baslik = f"📂 {exp} — {len(exp_df)} varlık"
            if eski > 0:
                baslik += f"  ⚠️ {eski} varlığın fiyatı 7+ gün eski"
            with st.expander(baslik):
                def renk_exp(row):
                    if row["Güncelleme"] > 30:
                        return ["background-color: #FFE0E0"] * len(row)
                    elif row["Güncelleme"] > 7:
                        return ["background-color: #FFF9C4"] * len(row)
                    return [""] * len(row)
                goster = exp_df[["Kod", "Ad", "Tür", "PB", "TWR % (TL)", "TWR % (PB)", "Yıllık (TL)", "Yıllık (PB)", "Son Fiyat", "Güncelleme"]].copy()
                st.dataframe(goster.style.apply(renk_exp, axis=1), use_container_width=True)

        # TWR açıklaması sayfanın üstünde yer almaktadır.

# ==========================================
# SAYFA 3: AYLIK ÖZET
# ==========================================
elif sayfa == "📅 Aylık Özet":
    st.title("📅 Aylık Portföy Özeti")
    st.markdown("---")

    yil = st.selectbox("Yıl seçin:", [2024, 2025, 2026, 2027], index=2)

    st.markdown("---")

    # ==========================================
    # AYLIK ÖZET TABLOSU
    # ==========================================
    with st.spinner("Hesaplanıyor..."):
        ozet_df = aylik_portfoy_ozeti(yil)

    if ozet_df.empty:
        st.info("Veri bulunamadı.")
    else:
        # --- USD tablosu: her ayın değerlerini o ayın kuruna böl ---
        usd_satirlar = []
        for _, row in ozet_df.iterrows():
            ay_str = row["Ay"]  # "2026-01"
            ay_basi_kur = kur_getir("USD", f"{ay_str}-01")
            # Ay sonu kuru: bir sonraki ayın 1'ine en yakın kur
            ay_no = int(ay_str.split("-")[1])
            if ay_no == 12:
                ay_sonu_tarih = f"{yil+1}-01-01"
            else:
                ay_sonu_tarih = f"{yil}-{str(ay_no+1).zfill(2)}-01"
            ay_sonu_kur = kur_getir("USD", ay_sonu_tarih)

            ay_basi_usd  = row["Ay Başı"]   / ay_basi_kur
            dis_giris_usd = row["Dış Giriş"] / ay_basi_kur if row["Dış Giriş"] else 0
            dis_cikis_usd = row["Dış Çıkış"] / ay_basi_kur if row["Dış Çıkış"] else 0
            ay_sonu_usd  = row["Ay Sonu"]   / ay_sonu_kur
            getiri_usd   = ay_sonu_usd - ay_basi_usd - dis_giris_usd + dis_cikis_usd

            usd_satirlar.append({
                "Ay"         : ay_str,
                "Ay Başı"   : round(ay_basi_usd, 2),
                "Dış Giriş" : round(dis_giris_usd, 2),
                "Dış Çıkış" : round(dis_cikis_usd, 2),
                "Getiri"     : round(getiri_usd, 2),
                "Ay Sonu"    : round(ay_sonu_usd, 2),
            })
        usd_ozet_df = pd.DataFrame(usd_satirlar)

        # ==========================================
        # GELECEKTEKİ AYLARI BOŞ GÖSTER
        # ==========================================
        # fiyat_gecmisi'ndeki en son tarih hangi ay?
        # O aydan sonraki satırların tüm sayısal sütunlarını None yap.
        son_fiyat_tarihi_ozet = sql_oku(
            "SELECT MAX(tarih) as t FROM fiyat_gecmisi", veritabani_baglan()
        ).iloc[0]["t"]

        sutunlar = ["Ay Başı", "Dış Giriş", "Dış Çıkış", "Getiri", "Ay Sonu"]

        if son_fiyat_tarihi_ozet:
            from datetime import datetime as _dt3
            son_dt_ozet = _dt3.strptime(son_fiyat_tarihi_ozet, "%Y-%m-%d")
            son_ay_ozet = son_dt_ozet.month   # örn: 6
            son_yil_ozet = son_dt_ozet.year   # örn: 2026

            if son_yil_ozet == yil:
                # Son ay dahil, sonrası boş
                for sutun in sutunlar:
                    ozet_df[sutun] = ozet_df.apply(
                        lambda r: r[sutun] if int(r["Ay"].split("-")[1]) <= son_ay_ozet else None, axis=1
                    )
                    usd_ozet_df[sutun] = usd_ozet_df.apply(
                        lambda r: r[sutun] if int(r["Ay"].split("-")[1]) <= son_ay_ozet else None, axis=1
                    )
            elif son_yil_ozet < yil:
                # Seçili yıl tamamen gelecek → tüm satırlar boş
                for sutun in sutunlar:
                    ozet_df[sutun] = None
                    usd_ozet_df[sutun] = None
            # son_yil_ozet > yil → seçili yıl tamamen geçmiş, hiçbir şey yapma

        # Metriklerde sadece dolu satırları kullan
        tl_dolu  = ozet_df.dropna(subset=["Ay Sonu"])
        usd_dolu = usd_ozet_df.dropna(subset=["Ay Sonu"])

        # ==========================================
        # METRİKLER (tabloların üstünde)
        # ==========================================
        # Son ay getirisi
        son_ay_getiri_tl  = tl_dolu["Getiri"].iloc[-1]  if not tl_dolu.empty else 0
        son_ay_getiri_usd = usd_dolu["Getiri"].iloc[-1] if not usd_dolu.empty else 0
        son_ay_adi        = tl_dolu["Ay"].iloc[-1]       if not tl_dolu.empty else "—"

        col1, col2, col3 = st.columns(3)
        col1.metric("Yılbaşından Bugüne Getiri", f"{tl_dolu['Getiri'].sum():,.0f} TL")
        col2.metric(f"Son Ay Getirisi ({son_ay_adi})", f"{son_ay_getiri_tl:,.0f} TL")
        col3.metric("Güncel Değer",              f"{tl_dolu['Ay Sonu'].iloc[-1]:,.0f} TL" if not tl_dolu.empty else "—")

        col4, col5, col6 = st.columns(3)
        col4.metric("Yılbaşından Bugüne Getiri (USD)", f"${usd_dolu['Getiri'].sum():,.0f}")
        col5.metric(f"Son Ay Getirisi ({son_ay_adi})", f"${son_ay_getiri_usd:,.0f}")
        col6.metric("Güncel Değer (USD)",           f"${usd_dolu['Ay Sonu'].iloc[-1]:,.0f}" if not usd_dolu.empty else "—")

        st.markdown("---")

        # --- Sekmeler: TL ve USD ---
        tab_tl, tab_usd = st.tabs(["🇹🇷 TL Bazında", "🇺🇸 USD Bazında"])

        with tab_tl:
            st.dataframe(
                ozet_df.style.format({
                    "Ay Başı"   : "{:,.0f}",
                    "Dış Giriş" : "{:,.0f}",
                    "Dış Çıkış" : "{:,.0f}",
                    "Getiri"    : "{:,.0f}",
                    "Ay Sonu"   : "{:,.0f}",
                }, na_rep=""),
                use_container_width=True
            )

        with tab_usd:
            st.dataframe(
                usd_ozet_df.style.format({
                    "Ay Başı"   : "${:,.0f}",
                    "Dış Giriş" : "${:,.0f}",
                    "Dış Çıkış" : "${:,.0f}",
                    "Getiri"    : "${:,.0f}",
                    "Ay Sonu"   : "${:,.0f}",
                }, na_rep=""),
                use_container_width=True
            )

        st.caption("ℹ️ Getiri = Ay Sonu Değer − Ay Başı Değer − Dış Giriş + Dış Çıkış. "
                   "Dış nakit akışları düzeltilmiştir, böylece portföye yeni para koymak getiri olarak sayılmaz. "
                   "USD versiyonunda her ayın değeri o ayın kendi USD kuru ile çevrilir.")

        # ==========================================
        # AYLIK GETİRİ GRAFİKLERİ (TL + USD)
        # ==========================================
        st.markdown("---")
        import plotly.express as px

        grafik_col_tl, grafik_col_usd = st.columns(2)

        with grafik_col_tl:
            fig_tl = px.bar(
                ozet_df,
                x="Ay",
                y="Getiri",
                title=f"{yil} Yılı — Aylık Getiri (TL)",
            )
            fig_tl.update_traces(marker_color="#1f77b4")
            st.plotly_chart(fig_tl, use_container_width=True)

        with grafik_col_usd:
            fig_usd = px.bar(
                usd_ozet_df,
                x="Ay",
                y="Getiri",
                title=f"{yil} Yılı — Aylık Getiri (USD)",
            )
            fig_usd.update_traces(marker_color="#2ca02c")
            st.plotly_chart(fig_usd, use_container_width=True)

        # ==========================================
        # DAĞILIM VERİSİNİ HESAPLA (ay detayı + matris için ortak)
        # ==========================================
        with st.spinner("Dağılım hesaplanıyor..."):
            dagilim_df = aylik_dagilim_hesapla(yil)

        # ==========================================
        # AY DETAYI — Tür → Exposure → Varlık kırılımı
        # ==========================================
        st.markdown("---")
        st.subheader("🔍 Ay Detayı")

        if dagilim_df.empty:
            st.info("Dağılım verisi bulunamadı.")
        else:
            # Ay seçici
            ay_secenekleri = {f"{i} — {AY_ISIMLERI[i-1]}": i for i in range(1, 13)}
            secilen_ay_str = st.selectbox("Ay seçin:", list(ay_secenekleri.keys()), index=min(date.today().month, 12) - 1)
            secilen_ay = ay_secenekleri[secilen_ay_str]
            secilen_ay_adi = AY_ISIMLERI[secilen_ay - 1]

            # Seçilen ayda değeri olan varlıkları filtrele
            ay_df = dagilim_df[dagilim_df[secilen_ay_adi] > 0].copy()

            if ay_df.empty:
                st.info(f"{secilen_ay_adi} ayında değeri olan varlık yok.")
            else:
                # USD kuru (o ayın sonu)
                if secilen_ay == 12:
                    ay_sonu_tarih_detay = f"{yil+1}-01-01"
                else:
                    ay_sonu_tarih_detay = f"{yil}-{str(secilen_ay+1).zfill(2)}-01"
                usd_kur_detay = kur_getir("USD", ay_sonu_tarih_detay)

                ay_df["Değer (TL)"] = ay_df[secilen_ay_adi]
                ay_df["Değer (USD)"] = ay_df[secilen_ay_adi] / usd_kur_detay

                # Toplam
                toplam_tl = ay_df["Değer (TL)"].sum()
                toplam_usd = ay_df["Değer (USD)"].sum()

                col_m1, col_m2 = st.columns(2)
                col_m1.metric(f"{secilen_ay_adi} Ay Sonu (TL)", f"{toplam_tl:,.0f} TL")
                col_m2.metric(f"{secilen_ay_adi} Ay Sonu (USD)", f"${toplam_usd:,.0f}")

                # --- Yardımcı fonksiyonlar ---
                def ay_deger_ozet(filtre_df):
                    return {
                        "Değer (TL)": filtre_df["Değer (TL)"].sum(),
                        "Değer (USD)": filtre_df["Değer (USD)"].sum(),
                    }

                def ay_deger_str(oz):
                    return (
                        f"Değer: **{oz['Değer (TL)']:,.0f} TL** | "
                        f"USD: **${oz['Değer (USD)']:,.0f}**"
                    )

                # --- Tür → Exposure → Varlık expander'ları ---
                turler = sorted(ay_df["Tür"].unique())

                for tur in turler:
                    tur_df = ay_df[ay_df["Tür"] == tur]
                    tur_oz = ay_deger_ozet(tur_df)

                    with st.expander(f"📂 {tur}  —  {ay_deger_str(tur_oz)}"):
                        # Exposure özet tablosu
                        exp_grp = tur_df.groupby("Exposure").agg(
                            DegerTL=("Değer (TL)", "sum"),
                            DegerUSD=("Değer (USD)", "sum"),
                        ).reset_index()
                        exp_grp.columns = ["Exposure", "Değer (TL)", "Değer (USD)"]
                        exp_grp["Yüzde"] = (exp_grp["Değer (TL)"] / toplam_tl * 100).round(1)
                        exp_grp["Yüzde"] = exp_grp["Yüzde"].apply(lambda x: f"%{x}")

                        st.markdown("**Exposure Özeti**")
                        st.dataframe(
                            exp_grp.style.format({
                                "Değer (TL)": "{:,.0f}",
                                "Değer (USD)": "${:,.0f}",
                            }),
                            use_container_width=True, hide_index=True
                        )

                        # Varlık listesi
                        st.markdown("**Varlık Listesi**")
                        goster = tur_df[["Kod", "Exposure", "PB", "Değer (TL)", "Değer (USD)"]].copy()
                        goster = goster.sort_values("Değer (TL)", ascending=False)
                        st.dataframe(
                            goster.style.format({
                                "Değer (TL)": "{:,.0f}",
                                "Değer (USD)": "${:,.0f}",
                            }),
                            use_container_width=True, hide_index=True
                        )

        # ==========================================
        # AYLIK VARLIK DAĞILIMI (matris)
        # ==========================================
        st.markdown("---")
        st.subheader("📊 Aylık Varlık Dağılımı (Ay Sonu Değerleri)")

        if dagilim_df.empty:
            st.info("Dağılım verisi bulunamadı.")
        else:
            # "Exposure — Tür" kategorisi oluştur
            dagilim_df["Kategori"] = dagilim_df["Exposure"] + " — " + dagilim_df["Tür"]

            # Exposure-Tür bazında grupla (ay sütunlarını topla)
            grup_df = dagilim_df.groupby("Kategori")[AY_ISIMLERI].sum()

            # Toplam satırı ekle (Aylık Özet'teki Ay Sonu ile eşleşir)
            grup_df.loc["TOPLAM"] = grup_df.sum()

            # ==========================================
            # GELECEKTEKİ AYLARI BOŞ GÖSTER
            # ==========================================
            # Son güncel fiyat tarihi hangi ay? O aydan sonrasını None yap.
            # Mantık: ay_sonu_tarihi = "bir sonraki ayın 1'i"
            # Eğer bu tarihte fiyat verisi yoksa (kur_getir hariç) → boş göster.
            # En basit yaklaşım: fiyat_gecmisi'ndeki MAX(tarih)'i bul,
            # hangi aya denk geldiğini hesapla, o ay dahil sonrasını boş bırak.
            son_fiyat_tarihi = sql_oku(
                "SELECT MAX(tarih) as t FROM fiyat_gecmisi", veritabani_baglan()
            ).iloc[0]["t"]

            if son_fiyat_tarihi:
                from datetime import datetime as _dt2
                son_dt = _dt2.strptime(son_fiyat_tarihi, "%Y-%m-%d")
                # Hangi ay sonu hesaplamasına girer? ay_sonu = "bir sonraki ayın 1'i"
                # Eğer son fiyat 15 Haziran ise:
                #   Haziran sonu = 2026-07-01 → bu tarihte veri var mı?
                #   Veri ancak 2026-06-15'e kadar var, 2026-07-01 bilinmiyor.
                # Bu yüzden: son fiyatın AYI = o ay dahil son dolu ay.
                # Temmuz ve sonrası (ay_no >= son_ay + 1) → boş.
                son_ay_no = son_dt.month  # örn: 6 (Haziran)
                son_yil   = son_dt.year

                # Sadece seçili yıl için geçerli
                if son_yil == yil:
                    for i, ay_adi in enumerate(AY_ISIMLERI):
                        ay_no = i + 1  # 1=Oca, 2=Şub, ...
                        if ay_no > son_ay_no:
                            # Gelecek ay → None yap (boş gösterilecek)
                            grup_df[ay_adi] = grup_df[ay_adi].apply(lambda x: None)
                elif son_yil < yil:
                    # Seçili yıl tamamen gelecekte → tüm aylar boş
                    for ay_adi in AY_ISIMLERI:
                        grup_df[ay_adi] = grup_df[ay_adi].apply(lambda x: None)
                # son_yil > yil: seçili yıl tamamen geçmişte → hiçbir şey yapma

            # --- USD versiyonu: her ayın değerini o ayın USD kuruna böl ---
            usd_grup_df = grup_df.copy()
            for i, ay_adi in enumerate(AY_ISIMLERI):
                ay_no = i + 1
                if ay_no == 12:
                    ay_sonu_tarih = f"{yil+1}-01-01"
                else:
                    ay_sonu_tarih = f"{yil}-{str(ay_no+1).zfill(2)}-01"
                usd_kur = kur_getir("USD", ay_sonu_tarih)
                # None olan hücreleri koruyarak böl
                usd_grup_df[ay_adi] = usd_grup_df[ay_adi].apply(
                    lambda x: x / usd_kur if x is not None and pd.notna(x) else None
                )

            # --- Sekmeler: TL ve USD ---
            dag_tab_tl, dag_tab_usd = st.tabs(["🇹🇷 TL Dağılım", "🇺🇸 USD Dağılım"])

            with dag_tab_tl:
                st.dataframe(
                    grup_df.style.format("{:,.0f}", na_rep=""),
                    use_container_width=True
                )

            with dag_tab_usd:
                st.dataframe(
                    usd_grup_df.style.format("${:,.0f}", na_rep=""),
                    use_container_width=True
                )

            # --- Varlık bazında detay (açılır) ---
            with st.expander("📋 Varlık Bazında Detay"):
                detay_df = dagilim_df[["Kod", "Tür", "Exposure", "PB"] + AY_ISIMLERI].copy()

                # Gelecek ayları detay tablosunda da boş yap
                if son_fiyat_tarihi and son_yil == yil:
                    for i, ay_adi in enumerate(AY_ISIMLERI):
                        if i + 1 > son_ay_no:
                            detay_df[ay_adi] = None
                elif son_fiyat_tarihi and son_yil < yil:
                    for ay_adi in AY_ISIMLERI:
                        detay_df[ay_adi] = None

                # USD versiyonu: her ayın değerini o ayın kuruna böl
                detay_usd_df = detay_df.copy()
                for i, ay_adi in enumerate(AY_ISIMLERI):
                    ay_no = i + 1
                    if ay_no == 12:
                        ay_sonu_tarih = f"{yil+1}-01-01"
                    else:
                        ay_sonu_tarih = f"{yil}-{str(ay_no+1).zfill(2)}-01"
                    usd_kur = kur_getir("USD", ay_sonu_tarih)
                    detay_usd_df[ay_adi] = detay_usd_df[ay_adi].apply(
                        lambda x: x / usd_kur if x is not None and pd.notna(x) else None
                    )

                detay_tab_tl, detay_tab_usd = st.tabs(["🇹🇷 TL", "🇺🇸 USD"])

                with detay_tab_tl:
                    st.dataframe(
                        detay_df.style.format(
                            {ay: "{:,.0f}" for ay in AY_ISIMLERI}, na_rep=""
                        ),
                        use_container_width=True
                    )

                with detay_tab_usd:
                    st.dataframe(
                        detay_usd_df.style.format(
                            {ay: "${:,.0f}" for ay in AY_ISIMLERI}, na_rep=""
                        ),
                        use_container_width=True
                    )

    # ==========================================
    # NAKİT AKIŞI GİRİŞ FORMU (sayfanın en altı)
    # ==========================================
    st.markdown("---")
    with st.expander("💸 Aylık Dış Giriş / Çıkış Girişi", expanded=False):
        st.caption("Portföye dışarıdan giren veya çıkan toplam nakit miktarını ay bazında girin.")

        baglanti = veritabani_baglan()
        mevcut_akislar = sql_oku("""
            SELECT ay, dis_giris, dis_cikis, notlar
            FROM portfoy_akislari
            WHERE yil = ?
            ORDER BY ay
        """, baglanti, params=(yil,))

        # Mevcut kayıtları göster (diptoplam satırıyla)
        if not mevcut_akislar.empty:
            # Diptoplam satırı ekle
            toplam_satir = pd.DataFrame([{
                "ay": "TOPLAM",
                "dis_giris": mevcut_akislar["dis_giris"].sum(),
                "dis_cikis": mevcut_akislar["dis_cikis"].sum(),
                "notlar": "",
            }])
            akislar_goster = pd.concat([mevcut_akislar, toplam_satir], ignore_index=True)

            st.dataframe(
                akislar_goster.style.format({
                    "dis_giris": "{:,.0f}",
                    "dis_cikis": "{:,.0f}",
                }),
                use_container_width=True, hide_index=True
            )

        # Yeni giriş formu
        with st.form("akis_formu"):
            col1, col2, col3 = st.columns(3)
            with col1:
                akis_ay = st.selectbox("Ay", list(range(1, 13)),
                    format_func=lambda x: f"{x} — {['', 'Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık'][x]}")
            with col2:
                akis_giris = st.number_input("Dış Giriş (TL)", min_value=0.0, step=1000.0, format="%.2f")
            with col3:
                akis_cikis = st.number_input("Dış Çıkış (TL)", min_value=0.0, step=1000.0, format="%.2f")
            akis_notlar = st.text_input("Notlar", placeholder="İsteğe bağlı")

            akis_kaydet = st.form_submit_button("💾 Kaydet")

            if akis_kaydet:
                baglanti = veritabani_baglan()
                cursor = baglanti.cursor()
                cursor.execute("""
                    INSERT OR REPLACE INTO portfoy_akislari (yil, ay, dis_giris, dis_cikis, notlar)
                    VALUES (?, ?, ?, ?, ?)
                """, (yil, akis_ay, akis_giris, akis_cikis, akis_notlar))
                baglanti.commit()
                senkronize_et()
                st.success(f"✅ {yil}-{str(akis_ay).zfill(2)} akışı kaydedildi!")
                import time
                time.sleep(1)
                st.rerun()

# ==========================================
# SAYFA: DÖNEMSEL KARŞILAŞTIRMA
# ==========================================
elif sayfa == "🔄 Dönemsel Karşılaştırma":
    st.title("🔄 Dönemsel Karşılaştırma")
    st.caption("İki tarih arasındaki portföy değeri değişimini kategori bazında karşılaştırır.")
    st.markdown("---")

    # --- Tarih seçiciler ---
    dk_col1, dk_col2 = st.columns(2)
    with dk_col1:
        dk_baslangic = st.date_input(
            "Başlangıç Tarihi",
            value=date.today().replace(day=1) - timedelta(days=1),  # geçen ay sonu
            key="dk_baslangic"
        )
    with dk_col2:
        dk_bitis = st.date_input(
            "Bitiş Tarihi",
            value=date.today(),
            key="dk_bitis"
        )

    if dk_baslangic >= dk_bitis:
        st.warning("Başlangıç tarihi, bitiş tarihinden önce olmalıdır.")
    else:
        bas_str = str(dk_baslangic)
        bit_str = str(dk_bitis)

        with st.spinner("Hesaplanıyor..."):
            dk_df = donemsel_karsilastirma_hesapla(bas_str, bit_str)

        if dk_df.empty:
            st.info("Seçilen tarihler için veri bulunamadı.")
        else:
            # --- "Exposure — Tür" kategorisi oluştur ---
            dk_df["Kategori"] = dk_df["Exposure"] + " — " + dk_df["Tür"]

            # --- Fark ve Fark % hesapla ---
            dk_df["Fark"] = dk_df["Bitiş"] - dk_df["Başlangıç"]
            dk_df["Fark %"] = dk_df.apply(
                lambda r: (r["Fark"] / r["Başlangıç"] * 100) if r["Başlangıç"] > 0 else None,
                axis=1
            )

            # --- Kategori bazında grupla ---
            dk_grup = dk_df.groupby("Kategori").agg(
                Başlangıç=("Başlangıç", "sum"),
                Bitiş=("Bitiş", "sum"),
            ).reset_index()
            dk_grup["Fark"] = dk_grup["Bitiş"] - dk_grup["Başlangıç"]
            dk_grup["Fark %"] = dk_grup.apply(
                lambda r: (r["Fark"] / r["Başlangıç"] * 100) if r["Başlangıç"] > 0 else None,
                axis=1
            )

            # --- TOPLAM satırı ---
            toplam_bas = dk_grup["Başlangıç"].sum()
            toplam_bit = dk_grup["Bitiş"].sum()
            toplam_fark = toplam_bit - toplam_bas
            toplam_fark_pct = (toplam_fark / toplam_bas * 100) if toplam_bas > 0 else None

            toplam_satir = pd.DataFrame([{
                "Kategori": "TOPLAM",
                "Başlangıç": toplam_bas,
                "Bitiş": toplam_bit,
                "Fark": toplam_fark,
                "Fark %": toplam_fark_pct,
            }])
            dk_grup = pd.concat([dk_grup, toplam_satir], ignore_index=True)

            # --- Sütun başlıklarını tarihle göster ---
            bas_etiket = dk_baslangic.strftime("%d.%m.%Y")
            bit_etiket = dk_bitis.strftime("%d.%m.%Y")
            dk_grup = dk_grup.rename(columns={
                "Başlangıç": bas_etiket,
                "Bitiş": bit_etiket,
            })

            # --- USD versiyonu ---
            usd_kur_bas = kur_getir("USD", bas_str)
            usd_kur_bit = kur_getir("USD", bit_str)

            dk_grup_usd = dk_grup.copy()
            dk_grup_usd[bas_etiket] = dk_grup_usd[bas_etiket] / usd_kur_bas
            dk_grup_usd[bit_etiket] = dk_grup_usd[bit_etiket] / usd_kur_bit
            dk_grup_usd["Fark"] = dk_grup_usd[bit_etiket] - dk_grup_usd[bas_etiket]
            dk_grup_usd["Fark %"] = dk_grup_usd.apply(
                lambda r: (r["Fark"] / r[bas_etiket] * 100) if r[bas_etiket] > 0 else None,
                axis=1
            )

            # --- Metrikler ---
            mcol1, mcol2, mcol3 = st.columns(3)
            mcol1.metric(f"Portföy ({bas_etiket})", f"{toplam_bas:,.0f} TL")
            mcol2.metric(f"Portföy ({bit_etiket})", f"{toplam_bit:,.0f} TL")
            fark_str = f"{toplam_fark:+,.0f} TL"
            pct_str = f"({toplam_fark_pct:+.1f}%)" if toplam_fark_pct is not None else ""
            mcol3.metric("Değişim", fark_str, delta=pct_str)

            st.markdown("---")

            # --- TL / USD sekmeleri ---
            dk_tab_tl, dk_tab_usd = st.tabs(["🇹🇷 TL Bazında", "🇺🇸 USD Bazında"])

            with dk_tab_tl:
                st.dataframe(
                    dk_grup.style.format({
                        bas_etiket: "{:,.0f}",
                        bit_etiket: "{:,.0f}",
                        "Fark": "{:+,.0f}",
                        "Fark %": "{:+.1f}%",
                    }, na_rep="—"),
                    use_container_width=True, hide_index=True
                )

            with dk_tab_usd:
                st.dataframe(
                    dk_grup_usd.style.format({
                        bas_etiket: "${:,.0f}",
                        bit_etiket: "${:,.0f}",
                        "Fark": "${:+,.0f}",
                        "Fark %": "{:+.1f}%",
                    }, na_rep="—"),
                    use_container_width=True, hide_index=True
                )

            # --- Varlık bazında detay ---
            with st.expander("📋 Varlık Bazında Detay"):
                detay = dk_df[["Kod", "Kategori", "PB", "Başlangıç", "Bitiş", "Fark", "Fark %"]].copy()
                detay = detay.sort_values("Fark", ascending=False)

                # USD versiyonu
                detay_usd = detay.copy()
                detay_usd["Başlangıç"] = detay_usd["Başlangıç"] / usd_kur_bas
                detay_usd["Bitiş"] = detay_usd["Bitiş"] / usd_kur_bit
                detay_usd["Fark"] = detay_usd["Bitiş"] - detay_usd["Başlangıç"]
                detay_usd["Fark %"] = detay_usd.apply(
                    lambda r: (r["Fark"] / r["Başlangıç"] * 100) if r["Başlangıç"] > 0 else None,
                    axis=1
                )

                # Sütun isimlerini tarihle göster
                detay = detay.rename(columns={"Başlangıç": bas_etiket, "Bitiş": bit_etiket})
                detay_usd = detay_usd.rename(columns={"Başlangıç": bas_etiket, "Bitiş": bit_etiket})

                detay_tl_tab, detay_usd_tab = st.tabs(["🇹🇷 TL", "🇺🇸 USD"])

                with detay_tl_tab:
                    st.dataframe(
                        detay.style.format({
                            bas_etiket: "{:,.0f}",
                            bit_etiket: "{:,.0f}",
                            "Fark": "{:+,.0f}",
                            "Fark %": "{:+.1f}%",
                        }, na_rep="—"),
                        use_container_width=True, hide_index=True
                    )

                with detay_usd_tab:
                    st.dataframe(
                        detay_usd.style.format({
                            bas_etiket: "${:,.0f}",
                            bit_etiket: "${:,.0f}",
                            "Fark": "${:+,.0f}",
                            "Fark %": "{:+.1f}%",
                        }, na_rep="—"),
                        use_container_width=True, hide_index=True
                    )

            st.caption("ℹ️ Her tarih için: Değer = Son Fiyat × Net Adet × Kur (TL). "
                       "USD versiyonunda her tarih kendi günündeki USD kuru ile çevrilir.")

# ==========================================
# SAYFA: YATIRIM FONLARI
# ==========================================
elif sayfa == "🏛️ Yatırım Fonları":
    st.title("🏛️ Yatırım Fonları")
    st.markdown("---")

    # ==========================================
    # VERİ HAZIRLAMA
    # ==========================================
    baglanti = veritabani_baglan()

    # Bugünkü kurlar
    usd_kuru_bugun = bugunun_kuru("USD")

    # Kurları kenar çubuğunda göster
    eur_kuru_bugun = bugunun_kuru("EUR")
    gbp_kuru_bugun = bugunun_kuru("GBP")
    st.sidebar.markdown("---")
    st.sidebar.markdown("**💱 Güncel Kurlar**")
    st.sidebar.markdown(f"USD/TRY: **{usd_kuru_bugun:.4f}**")
    st.sidebar.markdown(f"EUR/TRY: **{eur_kuru_bugun:.4f}**")
    st.sidebar.markdown(f"GBP/TRY: **{gbp_kuru_bugun:.4f}**")

    # Sadece Yatırım Fonu ve BES Fonu türlerini çek
    FON_TURLERI = ("Yatırım Fonu", "BES Fonu")

    fon_varliklar = sql_oku("""
        SELECT id, kod, ad, tur, para_birimi, exposure
        FROM varliklar
        WHERE tur IN ('Yatırım Fonu', 'BES Fonu')
    """, baglanti)

    if fon_varliklar.empty:
        st.info("Portföyde yatırım fonu veya BES fonu bulunmuyor.")
    else:
        # Net adet hesapla (pozisyonu olan fonlar)
        baglanti2 = veritabani_baglan()
        net_adetler = sql_oku("""
            SELECT varlik_id,
                   SUM(CASE WHEN islem_turu = 'Alış' THEN adet ELSE -adet END) AS net_adet
            FROM islemler
            WHERE islem_turu IN ('Alış', 'Satış')
              AND varlik_id IN (SELECT id FROM varliklar WHERE tur IN ('Yatırım Fonu', 'BES Fonu'))
            GROUP BY varlik_id
            HAVING net_adet > 0
        """, baglanti2)

        if net_adetler.empty:
            st.info("Aktif fon pozisyonu bulunmuyor.")
        else:
            # Son fiyatları çek (MAX(tarih) pattern — MAX(id) kullanma!)
            baglanti3 = veritabani_baglan()
            son_fiyatlar_fon = sql_oku("""
                SELECT f1.varlik_id, f1.fiyat, f1.tarih AS son_fiyat_tarihi
                FROM fiyat_gecmisi f1
                INNER JOIN (
                    SELECT varlik_id, MAX(tarih) AS son_tarih
                    FROM fiyat_gecmisi
                    GROUP BY varlik_id
                ) f2 ON f1.varlik_id = f2.varlik_id AND f1.tarih = f2.son_tarih
            """, baglanti3)

            # Portföy etiketi: her fon için en büyük pozisyona sahip etiketi bul
            baglanti4 = veritabani_baglan()
            etiket_bilgi = sql_oku("""
                SELECT varlik_id,
                       COALESCE(portfoy_etiketi, '') AS portfoy_etiketi,
                       SUM(CASE WHEN islem_turu = 'Alış' THEN adet ELSE -adet END) AS net
                FROM islemler
                WHERE islem_turu IN ('Alış', 'Satış')
                  AND varlik_id IN (SELECT id FROM varliklar WHERE tur IN ('Yatırım Fonu', 'BES Fonu'))
                GROUP BY varlik_id, portfoy_etiketi
                HAVING net > 0
            """, baglanti4)

            # Varlık → portföy etiketi eşlemesi (en büyük pozisyon)
            varlik_etiket = {}
            if not etiket_bilgi.empty:
                for vid in etiket_bilgi["varlik_id"].unique():
                    vid_df = etiket_bilgi[etiket_bilgi["varlik_id"] == vid]
                    en_buyuk = vid_df.loc[vid_df["net"].idxmax()]
                    et = en_buyuk["portfoy_etiketi"]
                    varlik_etiket[vid] = et if et else "Belirtilmemiş"

            # Fon özet tablosu oluştur
            from datetime import datetime as _dt_fon
            bugun_fon = date.today()

            fon_ozet = []
            for _, fon in fon_varliklar.iterrows():
                # Net adet
                adet_row = net_adetler[net_adetler["varlik_id"] == fon["id"]]
                if adet_row.empty:
                    continue
                net_adet = float(adet_row["net_adet"].values[0])

                # Son fiyat
                fiyat_row = son_fiyatlar_fon[son_fiyatlar_fon["varlik_id"] == fon["id"]]
                if fiyat_row.empty:
                    continue
                son_fiyat = float(fiyat_row["fiyat"].values[0])
                son_fiyat_tarihi = fiyat_row["son_fiyat_tarihi"].values[0]

                # Kur
                pb = fon["para_birimi"] if fon["para_birimi"] else "TRY"
                kur = bugunun_kuru(pb)

                # Güncel değer
                deger_tl = net_adet * son_fiyat * kur
                deger_usd = deger_tl / usd_kuru_bugun

                # Fiyat güncelliği (kaç gün eski)
                try:
                    son_tarih_dt = _dt_fon.strptime(son_fiyat_tarihi, "%Y-%m-%d").date()
                    gun_farki = (bugun_fon - son_tarih_dt).days
                except Exception:
                    gun_farki = 999

                fon_ozet.append({
                    "varlik_id"        : fon["id"],
                    "Kod"              : fon["kod"],
                    "Ad"               : fon["ad"],
                    "Tür"              : fon["tur"],
                    "Exposure"         : fon["exposure"] if fon["exposure"] else "—",
                    "Portföy Etiketi"  : varlik_etiket.get(fon["id"], "Belirtilmemiş"),
                    "Adet"             : net_adet,
                    "Birim Fiyat"      : son_fiyat,
                    "Değer (TL)"       : deger_tl,
                    "Değer (USD)"      : deger_usd,
                    "Son Fiyat Tarihi" : son_fiyat_tarihi,
                    "Güncelleme"       : gun_farki,
                })

            if not fon_ozet:
                st.info("Fiyat verisi olan aktif fon pozisyonu bulunmuyor.")
            else:
                fon_df = pd.DataFrame(fon_ozet)
                toplam_fon_tl  = fon_df["Değer (TL)"].sum()
                toplam_fon_usd = fon_df["Değer (USD)"].sum()
                fon_sayisi     = len(fon_df)

                # ==========================================
                # 1) METRİK KARTLARI
                # ==========================================
                mc1, mc2, mc3 = st.columns(3)
                mc1.metric("💰 Toplam Fon Değeri (TL)", f"{toplam_fon_tl:,.0f} TL")
                mc2.metric("💵 Toplam Fon Değeri (USD)", f"${toplam_fon_usd:,.0f}")
                mc3.metric("📊 Fon Sayısı", f"{fon_sayisi}")

                # ==========================================
                # 2) SON FİYAT TARİHİ UYARISI
                # ==========================================
                eski_fonlar = fon_df[fon_df["Güncelleme"] > 7]
                if not eski_fonlar.empty:
                    uyari_mesajlari = []
                    for _, ef in eski_fonlar.iterrows():
                        seviye = "🔴" if ef["Güncelleme"] > 30 else "🟡"
                        uyari_mesajlari.append(
                            f"{seviye} **{ef['Kod']}** — son fiyat: {ef['Son Fiyat Tarihi']} ({ef['Güncelleme']} gün eski)"
                        )
                    st.warning("⚠️ Fiyatı güncel olmayan fonlar:\n\n" + "\n\n".join(uyari_mesajlari))

                st.markdown("---")

                # ==========================================
                # 3) EXPOSURE BAZINDA TABLO (TL / USD sekmeli)
                # ==========================================
                st.subheader("Exposure Bazında Fon Dağılımı")

                tl_tab, usd_tab = st.tabs(["🇹🇷 TL", "🇺🇸 USD"])

                for sekme, deger_sutun, birim, toplam_deger in [
                    (tl_tab,  "Değer (TL)",  "TL",  toplam_fon_tl),
                    (usd_tab, "Değer (USD)", "USD", toplam_fon_usd),
                ]:
                    with sekme:
                        # --- Exposure bazında özet tablo ---
                        exp_grup = fon_df.groupby("Exposure").agg(
                            Bakiye=(deger_sutun, "sum"),
                            Fon_Sayisi=("Kod", "count"),
                        ).reset_index()

                        exp_grup["Pay %"] = (exp_grup["Bakiye"] / toplam_deger * 100).round(1)

                        # Toplam satırı ekle
                        toplam_satir = pd.DataFrame([{
                            "Exposure"  : "TOPLAM",
                            "Bakiye"    : toplam_deger,
                            "Fon_Sayisi": fon_sayisi,
                            "Pay %"     : 100.0,
                        }])
                        exp_tablo = pd.concat([exp_grup, toplam_satir], ignore_index=True)

                        if birim == "TL":
                            st.dataframe(
                                exp_tablo.rename(columns={
                                    "Bakiye": "Güncel Bakiye",
                                    "Fon_Sayisi": "Fon Sayısı",
                                }).style.format({
                                    "Güncel Bakiye": "{:,.0f} TL",
                                    "Pay %": "%{:.1f}",
                                }),
                                use_container_width=True, hide_index=True
                            )
                        else:
                            st.dataframe(
                                exp_tablo.rename(columns={
                                    "Bakiye": "Güncel Bakiye",
                                    "Fon_Sayisi": "Fon Sayısı",
                                }).style.format({
                                    "Güncel Bakiye": "${:,.0f}",
                                    "Pay %": "%{:.1f}",
                                }),
                                use_container_width=True, hide_index=True
                            )

                        # --- Exposure expand → Portföy Etiketi → Fon listesi ---
                        st.markdown("##### Detay Kırılımı")

                        exposurelar_fon = sorted(fon_df["Exposure"].unique())

                        for exp in exposurelar_fon:
                            exp_fonlar = fon_df[fon_df["Exposure"] == exp]
                            exp_toplam = exp_fonlar[deger_sutun].sum()
                            exp_pay    = (exp_toplam / toplam_deger * 100) if toplam_deger else 0

                            if birim == "TL":
                                exp_baslik = f"📂 {exp}  —  {exp_toplam:,.0f} TL  |  Pay: %{exp_pay:.1f}"
                            else:
                                exp_baslik = f"📂 {exp}  —  ${exp_toplam:,.0f}  |  Pay: %{exp_pay:.1f}"

                            with st.expander(exp_baslik):
                                # Portföy etiketi bazında alt kırılım
                                etiketler = sorted(exp_fonlar["Portföy Etiketi"].unique())

                                for etiket in etiketler:
                                    et_fonlar = exp_fonlar[exp_fonlar["Portföy Etiketi"] == etiket]
                                    et_toplam = et_fonlar[deger_sutun].sum()
                                    et_pay    = (et_toplam / toplam_deger * 100) if toplam_deger else 0

                                    if birim == "TL":
                                        et_baslik = f"🏷️ {etiket}  —  {et_toplam:,.0f} TL  |  Pay: %{et_pay:.1f}"
                                    else:
                                        et_baslik = f"🏷️ {etiket}  —  ${et_toplam:,.0f}  |  Pay: %{et_pay:.1f}"

                                    with st.expander(et_baslik):
                                        # Fon detay tablosu
                                        def renk_fon(row):
                                            if row["Güncelleme"] > 30:
                                                return ["background-color: #FFE0E0"] * len(row)
                                            elif row["Güncelleme"] > 7:
                                                return ["background-color: #FFF9C4"] * len(row)
                                            return [""] * len(row)

                                        goster_kolonlar = ["Kod", "Ad", "Adet", "Birim Fiyat",
                                                           deger_sutun, "Son Fiyat Tarihi", "Güncelleme"]
                                        goster_df = et_fonlar[goster_kolonlar].copy()
                                        goster_df = goster_df.sort_values(deger_sutun, ascending=False)

                                        # Pay % ekle (toplam fon portföyü içindeki)
                                        goster_df["Pay %"] = (goster_df[deger_sutun] / toplam_deger * 100).round(1)

                                        if birim == "TL":
                                            fmt = {
                                                "Adet"             : "{:,.4f}",
                                                "Birim Fiyat"      : "{:,.6f}",
                                                "Değer (TL)"       : "{:,.0f}",
                                                "Pay %"            : "%{:.1f}",
                                            }
                                        else:
                                            fmt = {
                                                "Adet"             : "{:,.4f}",
                                                "Birim Fiyat"      : "{:,.6f}",
                                                "Değer (USD)"      : "${:,.0f}",
                                                "Pay %"            : "%{:.1f}",
                                            }

                                        st.dataframe(
                                            goster_df.style
                                                .apply(renk_fon, axis=1)
                                                .format(fmt),
                                            use_container_width=True, hide_index=True
                                        )

                # ==========================================
                # 4) PORTFÖY ETİKETİ BAZINDA PIE CHART
                # ==========================================
                st.markdown("---")
                st.subheader("Portföy Etiketi Bazında Dağılım")

                import plotly.express as px

                pie_tl_tab, pie_usd_tab = st.tabs(["🇹🇷 TL", "🇺🇸 USD"])

                for pie_sekme, deger_kol, birim_adi in [
                    (pie_tl_tab,  "Değer (TL)",  "TL"),
                    (pie_usd_tab, "Değer (USD)", "USD"),
                ]:
                    with pie_sekme:
                        pie_data = fon_df.groupby("Portföy Etiketi").agg(
                            Toplam=(deger_kol, "sum")
                        ).reset_index()
                        pie_data = pie_data.sort_values("Toplam", ascending=False)

                        if birim_adi == "TL":
                            pie_data["Etiket"] = pie_data.apply(
                                lambda r: f"{r['Portföy Etiketi']}: {r['Toplam']:,.0f} TL", axis=1
                            )
                        else:
                            pie_data["Etiket"] = pie_data.apply(
                                lambda r: f"{r['Portföy Etiketi']}: ${r['Toplam']:,.0f}", axis=1
                            )

                        fig = px.pie(
                            pie_data,
                            values="Toplam",
                            names="Portföy Etiketi",
                            title=f"Fon Portföyü — Portföy Etiketi Dağılımı ({birim_adi})",
                            hover_data={"Etiket": True, "Toplam": False, "Portföy Etiketi": False},
                        )
                        fig.update_traces(
                            textposition="inside",
                            textinfo="label+percent",
                            hovertemplate="<b>%{label}</b><br>Tutar: %{customdata[0]}<extra></extra>",
                            customdata=pie_data[["Etiket"]].values,
                        )
                        fig.update_layout(
                            showlegend=True,
                            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                            height=500,
                        )
                        st.plotly_chart(fig, use_container_width=True)

                # ==========================================
                # 5) FON PERFORMANS ÖZETİ (TWR)
                # ==========================================
                st.markdown("---")
                st.subheader("Fon Performans Özeti")

                donem_fon = st.selectbox("Dönem seçin:", [
                    "bu_ay", "son_3_ay", "son_6_ay", "bu_yil", "tum_zamanlar"
                ], format_func=lambda x: {
                    "bu_ay"        : "Bu Ay",
                    "son_3_ay"     : "Son 3 Ay",
                    "son_6_ay"     : "Son 6 Ay",
                    "bu_yil"       : "Bu Yıl (YTD)",
                    "tum_zamanlar" : "Tüm Zamanlar"
                }[x], key="fon_donem")

                # Dönem tarihlerini hesapla
                bugun_perf = date.today()
                if donem_fon == "bu_ay":
                    bas_tarih = bugun_perf.replace(day=1).strftime("%Y-%m-%d")
                elif donem_fon == "son_3_ay":
                    from dateutil.relativedelta import relativedelta as _rd_fon
                    bas_tarih = (bugun_perf - _rd_fon(months=3)).strftime("%Y-%m-%d")
                elif donem_fon == "son_6_ay":
                    from dateutil.relativedelta import relativedelta as _rd_fon
                    bas_tarih = (bugun_perf - _rd_fon(months=6)).strftime("%Y-%m-%d")
                elif donem_fon == "bu_yil":
                    bas_tarih = bugun_perf.replace(month=1, day=1).strftime("%Y-%m-%d")
                else:
                    bas_tarih = "2000-01-01"
                bit_tarih = bugun_perf.strftime("%Y-%m-%d")

                # Her fon için TWR hesapla
                perf_sonuclar = []
                for _, fon_row in fon_df.iterrows():
                    sonuc = twr_hesapla(fon_row["varlik_id"], bas_tarih, bit_tarih)
                    if sonuc is not None:
                        from datetime import datetime as _dt_perf
                        bas_dt = _dt_perf.strptime(sonuc["ilk_tarih"], "%Y-%m-%d")
                        bit_dt = _dt_perf.strptime(sonuc["son_tarih"], "%Y-%m-%d")
                        gun_sayisi = (bit_dt - bas_dt).days

                        yillik_tl = yilliklandir(sonuc["twr_tl"], gun_sayisi) if gun_sayisi > 0 else None

                        perf_sonuclar.append({
                            "Kod"           : fon_row["Kod"],
                            "Ad"            : fon_row["Ad"],
                            "Exposure"      : fon_row["Exposure"],
                            "Portföy Etiketi": fon_row["Portföy Etiketi"],
                            "TWR % (TL)"    : f"{sonuc['twr_tl']:.2f}%",
                            "Yıllık (TL)"   : f"{yillik_tl:.2f}%" if yillik_tl is not None else "—",
                            "Son Fiyat"     : fon_row["Son Fiyat Tarihi"],
                            "Güncelleme"    : fon_row["Güncelleme"],
                            "TWR_sayi"      : sonuc["twr_tl"],
                        })

                if perf_sonuclar:
                    perf_df = pd.DataFrame(perf_sonuclar)

                    # Ortalama TWR metrikleri
                    ort_twr = perf_df["TWR_sayi"].mean()
                    st.metric("📊 Fonlar Ort. TWR (TL)", f"{ort_twr:.2f}%")

                    # Renklendirme fonksiyonu
                    def renk_perf(row):
                        if row["Güncelleme"] > 30:
                            return ["background-color: #FFE0E0"] * len(row)
                        elif row["Güncelleme"] > 7:
                            return ["background-color: #FFF9C4"] * len(row)
                        return [""] * len(row)

                    goster_perf = perf_df[["Kod", "Ad", "Exposure", "Portföy Etiketi",
                                           "TWR % (TL)", "Yıllık (TL)",
                                           "Son Fiyat", "Güncelleme"]].copy()
                    goster_perf = goster_perf.sort_values("TWR % (TL)", ascending=False)

                    st.dataframe(
                        goster_perf.style.apply(renk_perf, axis=1),
                        use_container_width=True, hide_index=True
                    )
                else:
                    st.info("Seçilen dönem için yeterli fiyat verisi yok. (En az 2 fiyat kaydı gerekli.)")

# ==========================================
# SAYFA 4: VARLIK EKLE
# ==========================================
elif sayfa == "➕ Varlık Ekle":
    st.title("➕ Yeni Varlık Ekle")
    st.markdown("---")

    with st.form("varlik_ekle_formu"):
        col1, col2 = st.columns(2)

        with col1:
            kod = st.text_input("Varlık Kodu", placeholder="Örn: GARAN, BTC, USD")
            ad  = st.text_input("Varlık Adı",  placeholder="Örn: Garanti Bankası")

        with col2:
            tur = st.selectbox("Varlık Türü", [
                "BIST Hisse", "Yabancı Hisse", "Kripto",
                "TL Mevduat", "YP Mevduat", "Yatırım Fonu",
                "BES Fonu", "VIOP", "Forex CFD", "Fiziki Maden"
            ])
            para_birimi = st.selectbox("Para Birimi", ["TRY", "USD", "EUR", "GBP"])
            exposure    = st.selectbox("Exposure (Risk Kategorisi)", [
                "TL", "YP", "Emtia", "Kripto", "Spekülatif"
            ])

        kaydet = st.form_submit_button("💾 Kaydet")

        if kaydet:
            if kod == "":
                st.error("Varlık kodu boş bırakılamaz!")
            else:
                try:
                    baglanti = veritabani_baglan()
                    cursor   = baglanti.cursor()
                    cursor.execute("""
                        INSERT OR IGNORE INTO varliklar (kod, ad, tur, para_birimi, exposure)
                        VALUES (?, ?, ?, ?, ?)
                    """, (kod.upper(), ad, tur, para_birimi, exposure))
                    baglanti.commit()
                    senkronize_et()
                    st.success(f"{kod.upper()} başarıyla eklendi!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Hata: {e}")

# ==========================================
# SAYFA 5: VARLIK DÜZENLE / SİL
# ==========================================
elif sayfa == "✏️ Varlık Düzenle":
    st.title("✏️ Varlık Düzenle / Sil")
    st.markdown("---")

    baglanti     = veritabani_baglan()
    varliklar_df = sql_oku("SELECT * FROM varliklar", baglanti)

    if varliklar_df.empty:
        st.info("Henüz varlık eklenmemiş.")
    else:
        varlik_secenekleri = {
            f"{row['kod']} — {row['ad']}": row['id']
            for _, row in varliklar_df.iterrows()
        }

        secilen   = st.selectbox("Düzenlenecek varlığı seçin:", list(varlik_secenekleri.keys()))
        varlik_id = varlik_secenekleri[secilen]
        mevcut    = varliklar_df[varliklar_df["id"] == varlik_id].iloc[0]

        st.markdown("---")
        st.subheader("Bilgileri Güncelle")

        tur_listesi = [
            "BIST Hisse", "Yabancı Hisse", "Kripto",
            "TL Mevduat", "YP Mevduat", "Yatırım Fonu",
            "BES Fonu", "VIOP", "Forex CFD", "Fiziki Maden"
        ]

        with st.form("varlik_duzenle_formu"):
            col1, col2 = st.columns(2)

            with col1:
                yeni_kod = st.text_input("Varlık Kodu", value=mevcut["kod"])
                yeni_ad  = st.text_input("Varlık Adı",  value=mevcut["ad"])

            with col2:
                yeni_tur = st.selectbox("Varlık Türü", tur_listesi,
                    index=tur_listesi.index(mevcut["tur"]) if mevcut["tur"] in tur_listesi else 0)

                yeni_para_birimi = st.selectbox("Para Birimi", ["TRY", "USD", "EUR", "GBP"],
                    index=["TRY", "USD", "EUR", "GBP"].index(mevcut["para_birimi"])
                    if mevcut["para_birimi"] in ["TRY", "USD", "EUR", "GBP"] else 0)

                exp_listesi  = ["TL", "YP", "Emtia", "Kripto", "Spekülatif"]
                yeni_exposure = st.selectbox("Exposure", exp_listesi,
                    index=exp_listesi.index(mevcut["exposure"])
                    if mevcut["exposure"] in exp_listesi else 0)

            guncelle = st.form_submit_button("💾 Güncelle")

            if guncelle:
                baglanti = veritabani_baglan()
                cursor   = baglanti.cursor()
                cursor.execute("""
                    UPDATE varliklar
                    SET kod = ?, ad = ?, tur = ?, para_birimi = ?, exposure = ?
                    WHERE id = ?
                """, (yeni_kod.upper(), yeni_ad, yeni_tur, yeni_para_birimi, yeni_exposure, varlik_id))
                baglanti.commit()
                senkronize_et()
                st.success(f"✅ {yeni_kod.upper()} güncellendi!")
                import time
                time.sleep(1)
                st.rerun()

        st.markdown("---")
        st.subheader("⚠️ Varlığı Sil")
        st.warning(f"**{mevcut['kod']}** varlığını silmek istiyorsanız aşağıdaki butona basın. Bu işlem geri alınamaz!")

        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("🗑️ Sil", type="primary"):
                baglanti = veritabani_baglan()
                cursor   = baglanti.cursor()
                cursor.execute("DELETE FROM islemler WHERE varlik_id = ?",    (varlik_id,))
                cursor.execute("DELETE FROM fiyat_gecmisi WHERE varlik_id = ?", (varlik_id,))
                cursor.execute("DELETE FROM varliklar WHERE id = ?",           (varlik_id,))
                baglanti.commit()
                senkronize_et()
                st.success(f"✅ {mevcut['kod']} silindi!")
                import time
                time.sleep(1)
                st.rerun()

# ==========================================
# SAYFA 6: İŞLEM EKLE
# ==========================================
elif sayfa == "💰 İşlem Ekle":
    st.title("💰 Yeni İşlem Ekle")
    st.markdown("---")

    baglanti     = veritabani_baglan()
    varliklar_df = sql_oku("SELECT id, kod, ad, tur FROM varliklar", baglanti)

    if varliklar_df.empty:
        st.warning("Önce varlık eklemeniz gerekiyor.")
    else:
        varlik_secenekleri = {
            f"{row['kod']} — {row['ad']}": row['id']
            for _, row in varliklar_df.iterrows()
        }

        # Seçili varlığın türünü form dışında al (dinamik UI için)
        secilen_varlik_adi = st.selectbox("Varlık", list(varlik_secenekleri.keys()), key="islem_varlik_sec")
        secilen_id         = varlik_secenekleri[secilen_varlik_adi]
        secilen_tur        = varliklar_df[varliklar_df["id"] == secilen_id]["tur"].values[0]
        is_mevduat_islem   = secilen_tur in MEVDUAT_TURLERI

        if is_mevduat_islem:
            st.info("💡 Mevduat türü varlık — birim fiyat otomatik **1** kabul edilir. "
                    "**Adet** alanına bakiye tutarını (TL veya YP) girin.")

        # ==========================================
        # YENİ ARACI KURUM / ETİKET EKLEME (form dışında)
        # ==========================================
        with st.expander("➕ Yeni Aracı Kurum veya Portföy Etiketi Ekle"):
            ek_col1, ek_col2 = st.columns(2)
            with ek_col1:
                yeni_araci_gir = st.text_input("Yeni aracı kurum adı:", key="yeni_araci_ekle")
                if st.button("✅ Aracı Kurum Ekle", key="araci_ekle_btn"):
                    if yeni_araci_gir.strip():
                        araci_kurum_kaydet(yeni_araci_gir.strip())
                        st.success(f"'{yeni_araci_gir.strip()}' eklendi!")
                        import time; time.sleep(1); st.rerun()
                    else:
                        st.error("Aracı kurum adı boş olamaz!")
            with ek_col2:
                yeni_etiket_gir = st.text_input("Yeni portföy etiketi:", key="yeni_etiket_ekle")
                if st.button("✅ Etiket Ekle", key="etiket_ekle_btn"):
                    if yeni_etiket_gir.strip():
                        portfoy_etiketi_kaydet(yeni_etiket_gir.strip())
                        st.success(f"'{yeni_etiket_gir.strip()}' eklendi!")
                        import time; time.sleep(1); st.rerun()
                    else:
                        st.error("Etiket adı boş olamaz!")

        # ==========================================
        # FORM (dropdown'lar dahil)
        # ==========================================
        ak_listesi = araci_kurum_listesi()
        pe_listesi = portfoy_etiketi_listesi()

        with st.form("islem_ekle_formu"):
            col1, col2 = st.columns(2)

            with col1:
                st.markdown(f"**Seçili Varlık:** {secilen_varlik_adi}")
                islem_turu = st.selectbox("İşlem Türü", [
                    "Alış", "Satış", "Temettü", "Faiz", "Komisyon", "Dış Giriş", "Dış Çıkış"
                ])
                tarih = st.date_input("İşlem Tarihi", value=date.today())

            with col2:
                adet = st.number_input("Adet / Miktar (Bakiye Tutarı)", min_value=0.0, step=0.01)

                if is_mevduat_islem:
                    fiyat = 1.0
                    st.markdown("**Birim Fiyat:** 1.00 *(mevduat — otomatik)*")
                else:
                    fiyat = st.number_input("Birim Fiyat", min_value=0.0, step=0.01)

                notlar = st.text_input("Notlar", placeholder="İsteğe bağlı")
                secilen_araci_kurum     = st.selectbox("Aracı Kurum", ak_listesi)
                secilen_portfoy_etiketi = st.selectbox("Portföy Etiketi", pe_listesi)

            tutar = adet * fiyat
            st.info(f"Tahmini Tutar: {tutar:,.2f}")

            kaydet = st.form_submit_button("💾 Kaydet")

            if kaydet:
                if adet == 0:
                    st.error("Adet/tutar sıfır olamaz!")
                elif not is_mevduat_islem and fiyat == 0:
                    st.error("Birim fiyat sıfır olamaz!")
                else:
                    try:
                        baglanti = veritabani_baglan()
                        cursor   = baglanti.cursor()
                        cursor.execute("""
                            INSERT INTO islemler
                                (varlik_id, tarih, islem_turu, adet, fiyat, tutar, notlar, araci_kurum, portfoy_etiketi)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (secilen_id, str(tarih), islem_turu, adet, fiyat, tutar,
                              notlar, secilen_araci_kurum, secilen_portfoy_etiketi))
                        baglanti.commit()
                        senkronize_et()
                        st.success(f"✅ İşlem kaydedildi! {secilen_varlik_adi} — {adet:,.2f} × {fiyat} = {tutar:,.2f}")
                        import time
                        time.sleep(2)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Hata: {e}")

# ==========================================
# SAYFA 7: İŞLEM DÜZENLE / SİL
# ==========================================
elif sayfa == "✏️ İşlem Düzenle":
    st.title("✏️ İşlem Düzenle / Sil")
    st.markdown("---")

    baglanti    = veritabani_baglan()
    islemler_df = sql_oku("""
        SELECT i.id, i.tarih, v.kod, v.ad, i.islem_turu, i.adet, i.fiyat, i.tutar,
               i.notlar, i.araci_kurum, i.portfoy_etiketi
        FROM islemler i
        JOIN varliklar v ON i.varlik_id = v.id
        ORDER BY i.tarih DESC
    """, baglanti)

    if islemler_df.empty:
        st.info("Henüz işlem girilmemiş.")
    else:
        islem_secenekleri = {
            f"{row['tarih']} — {row['kod']} — {row['islem_turu']} — {row['adet']} adet x {row['fiyat']}": row['id']
            for _, row in islemler_df.iterrows()
        }

        secilen  = st.selectbox("Düzenlenecek işlemi seçin:", list(islem_secenekleri.keys()))
        islem_id = islem_secenekleri[secilen]
        mevcut   = islemler_df[islemler_df["id"] == islem_id].iloc[0]

        st.markdown("---")
        st.subheader("Bilgileri Güncelle")

        islem_turu_listesi = ["Alış", "Satış", "Temettü", "Faiz", "Komisyon", "Dış Giriş", "Dış Çıkış"]

        # ==========================================
        # YENİ ARACI KURUM / ETİKET EKLEME (form dışında)
        # ==========================================
        with st.expander("➕ Yeni Aracı Kurum veya Portföy Etiketi Ekle"):
            ek_col1, ek_col2 = st.columns(2)
            with ek_col1:
                yeni_araci_d = st.text_input("Yeni aracı kurum adı:", key="yeni_araci_duzenle")
                if st.button("✅ Aracı Kurum Ekle", key="araci_ekle_duzenle_btn"):
                    if yeni_araci_d.strip():
                        araci_kurum_kaydet(yeni_araci_d.strip())
                        st.success(f"'{yeni_araci_d.strip()}' eklendi!")
                        import time; time.sleep(1); st.rerun()
                    else:
                        st.error("Aracı kurum adı boş olamaz!")
            with ek_col2:
                yeni_etiket_d = st.text_input("Yeni portföy etiketi:", key="yeni_etiket_duzenle")
                if st.button("✅ Etiket Ekle", key="etiket_ekle_duzenle_btn"):
                    if yeni_etiket_d.strip():
                        portfoy_etiketi_kaydet(yeni_etiket_d.strip())
                        st.success(f"'{yeni_etiket_d.strip()}' eklendi!")
                        import time; time.sleep(1); st.rerun()
                    else:
                        st.error("Etiket adı boş olamaz!")

        # ==========================================
        # FORM (dropdown'lar dahil)
        # ==========================================
        ak_listesi_d  = araci_kurum_listesi()
        pe_listesi_d  = portfoy_etiketi_listesi()

        mevcut_araci  = mevcut["araci_kurum"] if mevcut["araci_kurum"] else ""
        mevcut_etiket = mevcut["portfoy_etiketi"] if mevcut["portfoy_etiketi"] else ""

        # Mevcut değer listede yoksa (eski/silinen değer) → listeye ekle
        if mevcut_araci and mevcut_araci not in ak_listesi_d:
            ak_listesi_d.append(mevcut_araci)
        if mevcut_etiket and mevcut_etiket not in pe_listesi_d:
            pe_listesi_d.append(mevcut_etiket)

        ak_idx = ak_listesi_d.index(mevcut_araci) if mevcut_araci in ak_listesi_d else 0
        pe_idx = pe_listesi_d.index(mevcut_etiket) if mevcut_etiket in pe_listesi_d else 0

        with st.form("islem_duzenle_formu"):
            col1, col2 = st.columns(2)

            with col1:
                yeni_islem_turu = st.selectbox("İşlem Türü", islem_turu_listesi,
                    index=islem_turu_listesi.index(mevcut["islem_turu"])
                    if mevcut["islem_turu"] in islem_turu_listesi else 0)
                yeni_tarih = st.date_input("İşlem Tarihi", value=pd.to_datetime(mevcut["tarih"]))
                yeni_adet  = st.number_input("Adet / Miktar", min_value=0.0, step=0.01, value=float(mevcut["adet"]))
                yeni_fiyat = st.number_input("Birim Fiyat",   min_value=0.0, step=0.01, value=float(mevcut["fiyat"]))

            with col2:
                yeni_notlar = st.text_input("Notlar", value=mevcut["notlar"] if mevcut["notlar"] else "")
                yeni_araci_kurum = st.selectbox("Aracı Kurum", ak_listesi_d, index=ak_idx)
                yeni_portfoy_etiketi = st.selectbox("Portföy Etiketi", pe_listesi_d, index=pe_idx)

            yeni_tutar = yeni_adet * yeni_fiyat
            st.info(f"Tahmini Tutar: {yeni_tutar:,.2f}")

            guncelle = st.form_submit_button("💾 Güncelle")

            if guncelle:
                if yeni_adet == 0 or yeni_fiyat == 0:
                    st.error("Adet ve fiyat sıfır olamaz!")
                else:
                    baglanti = veritabani_baglan()
                    cursor   = baglanti.cursor()
                    cursor.execute("""
                        UPDATE islemler
                        SET tarih = ?, islem_turu = ?, adet = ?, fiyat = ?, tutar = ?,
                            notlar = ?, araci_kurum = ?, portfoy_etiketi = ?
                        WHERE id = ?
                    """, (str(yeni_tarih), yeni_islem_turu, yeni_adet, yeni_fiyat,
                          yeni_tutar, yeni_notlar, yeni_araci_kurum, yeni_portfoy_etiketi,
                          islem_id))
                    baglanti.commit()
                    senkronize_et()
                    st.success("✅ İşlem güncellendi!")
                    import time
                    time.sleep(1)
                    st.rerun()

        st.markdown("---")
        st.subheader("⚠️ İşlemi Sil")
        st.warning(f"**{mevcut['tarih']} — {mevcut['kod']} — {mevcut['islem_turu']}** işlemini silmek istiyorsanız aşağıdaki butona basın.")

        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button("🗑️ Sil", type="primary"):
                baglanti = veritabani_baglan()
                cursor   = baglanti.cursor()
                cursor.execute("DELETE FROM islemler WHERE id = ?", (islem_id,))
                baglanti.commit()
                senkronize_et()
                st.success("✅ İşlem silindi!")
                import time
                time.sleep(1)
                st.rerun()

# ==========================================
# SAYFA: İŞLEM YÜKLE (Excel'den Toplu)
# ==========================================
elif sayfa == "📤 İşlem Yükle":
    st.title("📤 Excel'den Toplu İşlem Yükle")
    st.caption("İşlemlerinizi Excel dosyasında hazırlayıp toplu olarak yükleyin.")
    st.markdown("---")

    # ==========================================
    # 1) ŞABLON İNDİR
    # ==========================================
    st.subheader("1️⃣ Şablon İndir")
    st.info("Aşağıdaki butona basarak boş şablonu indirin, doldurun, sonra yükleyin.")

    from openpyxl import Workbook as _Wb
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
    from io import BytesIO as _BytesIO

    def sablon_olustur():
        """Sütun başlıkları ve örnek satır içeren Excel şablonu oluşturur."""
        wb = _Wb()
        ws = wb.active
        ws.title = "İşlemler"

        # --- Başlıklar ---
        basliklar = ["kod", "tarih", "islem_turu", "adet", "fiyat", "notlar", "araci_kurum", "portfoy_etiketi"]
        baslik_font = _Font(bold=True, color="FFFFFF", size=11)
        baslik_fill = _Fill("solid", fgColor="4472C4")
        baslik_align = _Align(horizontal="center")

        for col_no, baslik in enumerate(basliklar, 1):
            hucre = ws.cell(row=1, column=col_no, value=baslik)
            hucre.font = baslik_font
            hucre.fill = baslik_fill
            hucre.alignment = baslik_align

        # --- Örnek satır ---
        ornek = ["AAPL", "2025-06-15", "Alış", 10, 235.50, "İlk alım", "İş Yatırım", "Yatırım"]
        ornek_font = _Font(color="808080", italic=True)
        for col_no, deger in enumerate(ornek, 1):
            hucre = ws.cell(row=2, column=col_no, value=deger)
            hucre.font = ornek_font

        # --- Sütun genişlikleri ---
        genislikler = [14, 14, 14, 14, 14, 20, 18, 18]
        for i, g in enumerate(genislikler, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = g

        # --- Açıklama sayfası ---
        ws2 = wb.create_sheet("Açıklama")
        aciklamalar = [
            ["Sütun", "Zorunlu?", "Açıklama", "Örnek Değerler"],
            ["kod", "EVET", "Varlık kodu (varliklar tablosundaki kod)", "AAPL, GARAN, BTC"],
            ["tarih", "EVET", "İşlem tarihi (YYYY-MM-DD formatında)", "2025-06-15"],
            ["islem_turu", "EVET", "İşlem türü", "Alış, Satış, Temettü, Faiz, Komisyon"],
            ["adet", "EVET", "Miktar (mevduatta bakiye tutarı)", "10, 0.5, 50000"],
            ["fiyat", "EVET", "Birim fiyat (mevduatta 1 yazın)", "235.50, 1, 0.0045"],
            ["notlar", "HAYIR", "Açıklama (boş bırakılabilir)", "İlk alım, Temettü ödemesi"],
            ["araci_kurum", "HAYIR", "Aracı kurum adı (boş bırakılabilir)", "İş Yatırım, Akbank"],
            ["portfoy_etiketi", "HAYIR", "Portföy etiketi (boş bırakılabilir)", "Yatırım, Defans"],
        ]
        for satir_no, satir in enumerate(aciklamalar, 1):
            for col_no, deger in enumerate(satir, 1):
                hucre = ws2.cell(row=satir_no, column=col_no, value=deger)
                if satir_no == 1:
                    hucre.font = baslik_font
                    hucre.fill = baslik_fill
                    hucre.alignment = baslik_align
        for i, g in enumerate([12, 12, 45, 35], 1):
            ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = g

        buf = _BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    st.download_button(
        label="📥 Boş Şablonu İndir (.xlsx)",
        data=sablon_olustur(),
        file_name="islem_sablonu.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ==========================================
    # 2) DOSYA YÜKLE
    # ==========================================
    st.markdown("---")
    st.subheader("2️⃣ Dosya Yükle ve Kontrol Et")

    yuklenen = st.file_uploader("Excel dosyanızı seçin (.xlsx)", type=["xlsx"], key="islem_yukle")

    if yuklenen is not None:
        try:
            yukle_df = pd.read_excel(yuklenen, dtype={"kod": str, "notlar": str,
                                                       "araci_kurum": str, "portfoy_etiketi": str})
        except Exception as e:
            st.error(f"Dosya okunamadı: {e}")
            yukle_df = pd.DataFrame()

        if not yukle_df.empty:
            # --- Sütun adı kontrolü ---
            zorunlu_sutunlar = ["kod", "tarih", "islem_turu", "adet", "fiyat"]
            eksik_sutunlar = [s for s in zorunlu_sutunlar if s not in yukle_df.columns]

            if eksik_sutunlar:
                st.error(f"Eksik sütunlar: **{', '.join(eksik_sutunlar)}**. Şablondaki sütun adlarını kontrol edin.")
            else:
                # --- Boş satırları temizle ---
                yukle_df = yukle_df.dropna(subset=["kod", "tarih", "islem_turu", "adet", "fiyat"])

                if yukle_df.empty:
                    st.warning("Dosyada geçerli veri satırı bulunamadı.")
                else:
                    # --- Opsiyonel sütunları ekle (yoksa) ---
                    for opsiyonel in ["notlar", "araci_kurum", "portfoy_etiketi"]:
                        if opsiyonel not in yukle_df.columns:
                            yukle_df[opsiyonel] = ""
                    yukle_df["notlar"]           = yukle_df["notlar"].fillna("").astype(str)
                    yukle_df["araci_kurum"]      = yukle_df["araci_kurum"].fillna("").astype(str)
                    yukle_df["portfoy_etiketi"]  = yukle_df["portfoy_etiketi"].fillna("").astype(str)

                    # --- Tarih formatı düzelt ---
                    try:
                        yukle_df["tarih"] = pd.to_datetime(yukle_df["tarih"]).dt.strftime("%Y-%m-%d")
                    except Exception:
                        st.error("Tarih sütunundaki değerler okunamadı. YYYY-MM-DD formatında olduğundan emin olun.")
                        st.stop()

                    # --- Kod büyük harfe ---
                    yukle_df["kod"] = yukle_df["kod"].str.strip().str.upper()

                    # ==========================================
                    # 3) DOĞRULAMA
                    # ==========================================
                    baglanti = veritabani_baglan()
                    varliklar_db = sql_oku("SELECT id, kod, tur FROM varliklar", baglanti)
                    kod_id_map = dict(zip(varliklar_db["kod"], varliklar_db["id"]))
                    kod_tur_map = dict(zip(varliklar_db["kod"], varliklar_db["tur"]))

                    gecerli_turler = {"Alış", "Satış", "Temettü", "Faiz", "Komisyon", "Dış Giriş", "Dış Çıkış"}

                    hatalar = []
                    gecerli_satirlar = []

                    for idx, row in yukle_df.iterrows():
                        satir_no = idx + 2  # Excel'de 1. satır başlık, 2. satırdan veri başlar
                        satir_hatalari = []

                        # Kod kontrolü
                        if row["kod"] not in kod_id_map:
                            satir_hatalari.append(f"'{row['kod']}' kodu veritabanında bulunamadı")

                        # İşlem türü kontrolü
                        if row["islem_turu"] not in gecerli_turler:
                            satir_hatalari.append(f"Geçersiz işlem türü: '{row['islem_turu']}'")

                        # Sayısal alan kontrolü
                        try:
                            adet_val = float(row["adet"])
                            if adet_val <= 0:
                                satir_hatalari.append("Adet sıfır veya negatif olamaz")
                        except (ValueError, TypeError):
                            satir_hatalari.append(f"Adet sayısal değil: '{row['adet']}'")

                        try:
                            fiyat_val = float(row["fiyat"])
                            if fiyat_val < 0:
                                satir_hatalari.append("Fiyat negatif olamaz")
                        except (ValueError, TypeError):
                            satir_hatalari.append(f"Fiyat sayısal değil: '{row['fiyat']}'")

                        if satir_hatalari:
                            hatalar.append({"Satır": satir_no, "Kod": row["kod"],
                                            "Hata": " | ".join(satir_hatalari)})
                        else:
                            # Mevduat kontrolü: fiyatı otomatik 1 yap
                            varlik_tur = kod_tur_map.get(row["kod"], "")
                            if varlik_tur in MEVDUAT_TURLERI:
                                fiyat_val = 1.0

                            gecerli_satirlar.append({
                                "varlik_id"       : kod_id_map[row["kod"]],
                                "kod"             : row["kod"],
                                "tarih"           : row["tarih"],
                                "islem_turu"      : row["islem_turu"],
                                "adet"            : float(row["adet"]),
                                "fiyat"           : fiyat_val,
                                "tutar"           : round(float(row["adet"]) * fiyat_val, 4),
                                "notlar"          : row["notlar"],
                                "araci_kurum"     : row["araci_kurum"],
                                "portfoy_etiketi" : row["portfoy_etiketi"],
                            })

                    # ==========================================
                    # 4) SONUÇLARI GÖSTER
                    # ==========================================
                    st.markdown("---")
                    st.subheader("3️⃣ Doğrulama Sonucu")

                    toplam = len(yukle_df)
                    gecerli = len(gecerli_satirlar)
                    hatali = len(hatalar)

                    r_col1, r_col2, r_col3 = st.columns(3)
                    r_col1.metric("Toplam Satır", toplam)
                    r_col2.metric("Geçerli", gecerli)
                    r_col3.metric("Hatalı", hatali)

                    # --- Hataları göster ---
                    if hatalar:
                        st.error(f"⚠️ {hatali} satırda hata bulundu:")
                        hata_df = pd.DataFrame(hatalar)
                        st.dataframe(hata_df, use_container_width=True, hide_index=True)

                    # --- Geçerli satırları önizle ---
                    if gecerli_satirlar:
                        st.success(f"✅ {gecerli} satır yüklenmeye hazır:")
                        onizleme_df = pd.DataFrame(gecerli_satirlar)
                        st.dataframe(
                            onizleme_df[["kod", "tarih", "islem_turu", "adet", "fiyat", "tutar",
                                         "notlar", "araci_kurum", "portfoy_etiketi"]].style.format({
                                "adet": "{:,.4f}",
                                "fiyat": "{:,.4f}",
                                "tutar": "{:,.4f}",
                            }),
                            use_container_width=True, hide_index=True
                        )

                        # ==========================================
                        # 5) KAYDET
                        # ==========================================
                        st.markdown("---")
                        if hatalar:
                            st.warning("⚠️ Hatalı satırlar atlanacak, sadece geçerli satırlar kaydedilecek.")

                        if st.button("💾 Geçerli İşlemleri Kaydet", type="primary"):
                            import db as _db_mod
                            _db_mod._baglanti = None
                            baglanti = veritabani_baglan()
                            cursor = baglanti.cursor()

                            eklenen = 0
                            # Turso timeout önlemi: 20'li chunk'lar
                            chunk_boyut = 20

                            for i, satir in enumerate(gecerli_satirlar):
                                cursor.execute("""
                                    INSERT INTO islemler
                                        (varlik_id, tarih, islem_turu, adet, fiyat, tutar,
                                         notlar, araci_kurum, portfoy_etiketi)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (
                                    satir["varlik_id"], satir["tarih"], satir["islem_turu"],
                                    satir["adet"], satir["fiyat"], satir["tutar"],
                                    satir["notlar"], satir["araci_kurum"], satir["portfoy_etiketi"],
                                ))
                                eklenen += 1

                                # Chunk commit
                                if (i + 1) % chunk_boyut == 0:
                                    baglanti.commit()
                                    senkronize_et()
                                    _db_mod._baglanti = None
                                    baglanti = veritabani_baglan()
                                    cursor = baglanti.cursor()

                            # Kalan kayıtları commit et
                            baglanti.commit()
                            senkronize_et()
                            _db_mod._baglanti = None

                            st.success(f"✅ {eklenen} işlem başarıyla kaydedildi!")
                            import time
                            time.sleep(2)
                            st.rerun()
                    else:
                        st.warning("Geçerli satır bulunamadı. Lütfen hataları düzeltip tekrar yükleyin.")

# ==========================================
# SAYFA 8: İŞLEM GEÇMİŞİ
# ==========================================
elif sayfa == "📋 İşlem Geçmişi":
    st.title("📋 İşlem Geçmişi")
    st.markdown("---")

    baglanti = veritabani_baglan()
    df = sql_oku("""
        SELECT i.tarih, v.kod, v.ad, i.islem_turu, i.adet, i.fiyat, i.tutar,
               i.araci_kurum, i.portfoy_etiketi, i.notlar
        FROM islemler i
        JOIN varliklar v ON i.varlik_id = v.id
        ORDER BY i.tarih DESC
    """, baglanti)

    if df.empty:
        st.info("Henüz işlem girilmemiş.")
    else:
        # ==========================================
        # FİLTRELER
        # ==========================================
        with st.expander("🔍 Filtrele", expanded=False):
            f_col1, f_col2, f_col3, f_col4 = st.columns(4)

            with f_col1:
                varlik_secenekleri_f = ["Tümü"] + sorted(df["kod"].unique().tolist())
                filtre_varlik = st.multiselect("Varlık", varlik_secenekleri_f, default=["Tümü"], key="f_varlik")

            with f_col2:
                islem_turu_secenekleri = ["Tümü"] + sorted(df["islem_turu"].unique().tolist())
                filtre_islem_turu = st.multiselect("İşlem Türü", islem_turu_secenekleri, default=["Tümü"], key="f_islem_turu")

            with f_col3:
                ak_secenekleri = ["Tümü"] + sorted([x for x in df["araci_kurum"].unique().tolist() if x])
                filtre_ak = st.multiselect("Aracı Kurum", ak_secenekleri, default=["Tümü"], key="f_ak")

            with f_col4:
                tarih_baslangic = st.date_input("Başlangıç", value=pd.to_datetime(df["tarih"].min()), key="f_tarih_bas")
                tarih_bitis     = st.date_input("Bitiş", value=pd.to_datetime(df["tarih"].max()), key="f_tarih_bit")

        # Filtreleri uygula
        filtreli_df = df.copy()

        if "Tümü" not in filtre_varlik:
            filtreli_df = filtreli_df[filtreli_df["kod"].isin(filtre_varlik)]

        if "Tümü" not in filtre_islem_turu:
            filtreli_df = filtreli_df[filtreli_df["islem_turu"].isin(filtre_islem_turu)]

        if "Tümü" not in filtre_ak:
            filtreli_df = filtreli_df[filtreli_df["araci_kurum"].isin(filtre_ak)]

        filtreli_df = filtreli_df[
            (filtreli_df["tarih"] >= str(tarih_baslangic)) &
            (filtreli_df["tarih"] <= str(tarih_bitis))
        ]

        st.caption(f"Toplam {len(filtreli_df)} işlem gösteriliyor (toplam: {len(df)})")

        # ==========================================
        # TABLO
        # ==========================================
        st.dataframe(
            filtreli_df.style.format({
                "adet"  : "{:,.4f}",
                "fiyat" : "{:,.4f}",
                "tutar" : "{:,.2f}",
            }),
            use_container_width=True, hide_index=True
        )

        # ==========================================
        # EXCEL EXPORT
        # ==========================================
        import io
        buffer = io.BytesIO()
        filtreli_df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)

        st.download_button(
            label="📥 Excel'e Aktar",
            data=buffer,
            file_name="islem_gecmisi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ==========================================
# SAYFA 9: FİYAT GEÇMİŞİ
# ==========================================
elif sayfa == "🗓️ Fiyat Geçmişi":
    st.title("🗓️ Fiyat Geçmişi")
    st.markdown("---")

    baglanti     = veritabani_baglan()
    varliklar_df = sql_oku("SELECT id, kod, ad FROM varliklar ORDER BY kod", baglanti)

    if varliklar_df.empty:
        st.info("Henüz varlık eklenmemiş.")
    else:
        varlik_secenekleri = {
            f"{row['kod']} — {row['ad']}": row['id']
            for _, row in varliklar_df.iterrows()
        }

        secilen   = st.selectbox("Varlık seçin:", list(varlik_secenekleri.keys()))
        varlik_id = varlik_secenekleri[secilen]

        st.markdown("---")

        baglanti = veritabani_baglan()
        fiyat_df = sql_oku("""
            SELECT tarih, fiyat, kaynak
            FROM fiyat_gecmisi
            WHERE varlik_id = ?
            ORDER BY tarih DESC
        """, baglanti, params=(varlik_id,))

        col1, col2 = st.columns([2, 1])

        with col1:
            st.subheader("Kayıtlı Fiyatlar")
            if fiyat_df.empty:
                st.info("Bu varlık için fiyat kaydı yok.")
            else:
                st.dataframe(fiyat_df, use_container_width=True)
                st.caption(f"Toplam {len(fiyat_df)} kayıt")

        with col2:
            st.subheader("Fiyat Ekle")
            with st.form("fiyat_ekle_formu"):
                yeni_tarih = st.date_input("Tarih", value=date.today())
                yeni_fiyat = st.number_input("Fiyat", min_value=0.0, step=0.01)
                kaynak     = st.selectbox("Kaynak", ["manuel", "tefas", "yahoo", "tcmb"])
                ekle       = st.form_submit_button("➕ Ekle")

                if ekle:
                    if yeni_fiyat == 0:
                        st.error("Fiyat sıfır olamaz!")
                    else:
                        try:
                            baglanti = veritabani_baglan()
                            cursor   = baglanti.cursor()
                            cursor.execute("""
                                INSERT OR REPLACE INTO fiyat_gecmisi
                                    (varlik_id, tarih, fiyat, kaynak)
                                VALUES (?, ?, ?, ?)
                            """, (varlik_id, str(yeni_tarih), yeni_fiyat, kaynak))
                            baglanti.commit()
                            senkronize_et()
                            st.success("✅ Fiyat eklendi!")
                            import time
                            time.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Hata: {e}")

            st.markdown("---")
            st.subheader("Fiyat Sil")
            if not fiyat_df.empty:
                silinecek_tarih = st.selectbox("Silinecek tarih:", fiyat_df["tarih"].tolist())
                if st.button("🗑️ Sil", type="primary"):
                    baglanti = veritabani_baglan()
                    cursor   = baglanti.cursor()
                    cursor.execute("""
                        DELETE FROM fiyat_gecmisi
                        WHERE varlik_id = ? AND tarih = ?
                    """, (varlik_id, silinecek_tarih))
                    baglanti.commit()
                    senkronize_et()
                    st.success("✅ Fiyat silindi!")
                    import time
                    time.sleep(1)
                    st.rerun()

        if not fiyat_df.empty:
            st.markdown("---")
            st.subheader("📈 Fiyat Grafiği")
            import plotly.express as px
            grafik_df = fiyat_df.sort_values("tarih")
            fig = px.line(grafik_df, x="tarih", y="fiyat",
                         title=f"{secilen} Fiyat Geçmişi", markers=True)
            st.plotly_chart(fig, use_container_width=True)

# ==========================================
# SAYFA: FİYAT GÜNCELLE
# ==========================================
elif sayfa == "💱 Fiyat Güncelle":
    st.title("💱 Fiyat Güncelle")
    st.markdown("---")

    baglanti = veritabani_baglan()
    df = sql_oku("""
        SELECT
            v.id, v.kod, v.ad, v.tur,
            SUM(CASE WHEN i.islem_turu = 'Alış' THEN i.adet ELSE -i.adet END) AS toplam_adet
        FROM varliklar v
        LEFT JOIN islemler i ON v.id = i.varlik_id
        GROUP BY v.id
        HAVING toplam_adet > 0
    """, baglanti)

    son_fiyatlar = sql_oku("""
        SELECT f1.varlik_id, f1.fiyat, f1.tarih
        FROM fiyat_gecmisi f1
        INNER JOIN (
            SELECT varlik_id, MAX(tarih) AS son_tarih
            FROM fiyat_gecmisi
            GROUP BY varlik_id
        ) f2 ON f1.varlik_id = f2.varlik_id AND f1.tarih = f2.son_tarih
    """, baglanti)

    if df.empty:
        st.info("Henüz varlık eklenmemiş.")
    else:
        # --- Fiyat geçmişi eksik varlık uyarısı ---
        eksik_fiyat_fg = sql_oku("""
            SELECT
                v.kod, v.tur,
                MIN(i.tarih) AS ilk_islem,
                (SELECT MIN(f.tarih) FROM fiyat_gecmisi f WHERE f.varlik_id = v.id) AS ilk_fiyat
            FROM varliklar v
            JOIN islemler i ON v.id = i.varlik_id AND i.islem_turu = 'Alış'
            JOIN (
                SELECT varlik_id, SUM(CASE WHEN islem_turu = 'Alış' THEN adet ELSE -adet END) AS net
                FROM islemler WHERE islem_turu IN ('Alış', 'Satış')
                GROUP BY varlik_id HAVING net > 0
            ) p ON v.id = p.varlik_id
            WHERE v.tur NOT IN ('TL Mevduat', 'YP Mevduat')
            GROUP BY v.id
        """, veritabani_baglan())

        if not eksik_fiyat_fg.empty:
            uyarilar_fg = []
            for _, r in eksik_fiyat_fg.iterrows():
                if r["ilk_fiyat"] is None:
                    uyarilar_fg.append(f"**{r['kod']}** ({r['tur']}) — ilk işlem: {r['ilk_islem']}, fiyat verisi yok")
                else:
                    from datetime import datetime as _dt
                    islem_dt = _dt.strptime(r["ilk_islem"], "%Y-%m-%d")
                    fiyat_dt = _dt.strptime(r["ilk_fiyat"], "%Y-%m-%d")
                    bosluk = (fiyat_dt - islem_dt).days
                    if bosluk > 3:
                        uyarilar_fg.append(f"**{r['kod']}** ({r['tur']}) — ilk işlem: {r['ilk_islem']}, ilk fiyat: {r['ilk_fiyat']} ({bosluk} gün boşluk)")
            if uyarilar_fg:
                st.warning("⚠️ Fiyat verisi eksik:\n\n" + "\n\n".join(uyarilar_fg) + "\n\nAşağıdaki Geçmiş Veri Tamamla bölümünden çekin.")

        # ==========================================
        # OTOMATİK FİYAT ÇEKME BUTONLARI
        # ==========================================
        st.subheader("🤖 Otomatik Fiyat Çek")
        st.caption("Yahoo Finance'tan güncel fiyatları otomatik olarak çeker ve kaydeder.")

        # --- Satır 1: Yabancı Hisse, BIST Hisse, Kripto, Altın, Döviz Kuru ---
        oto_col1, oto_col2, oto_col3, oto_col4, oto_col5 = st.columns(5)

        with oto_col1:
            if st.button("📈 Yabancı Hisse Fiyatları", use_container_width=True):
                with st.spinner("Yahoo Finance'tan çekiliyor..."):
                    sayi = hisse_fiyatlari_cek()
                if sayi > 0:
                    st.success(f"✅ {sayi} hisse fiyatı güncellendi!")
                    import time
                    time.sleep(1)
                    st.rerun()
                else:
                    st.warning("Fiyat çekilemedi (piyasa kapalı olabilir).")

        with oto_col2:
            if st.button("🏦 BIST Hisse Fiyatları", use_container_width=True):
                with st.spinner("Yahoo Finance'tan BIST fiyatları çekiliyor..."):
                    sayi = bist_fiyatlari_cek()
                if sayi > 0:
                    st.success(f"✅ {sayi} BIST hisse fiyatı güncellendi!")
                    import time
                    time.sleep(1)
                    st.rerun()
                else:
                    st.warning("BIST fiyatı çekilemedi (piyasa kapalı olabilir veya BIST Hisse varlığı yok).")

        with oto_col3:
            if st.button("🪙 Kripto Fiyatları", use_container_width=True):
                with st.spinner("Yahoo Finance'tan kripto fiyatları çekiliyor..."):
                    sayi = kripto_fiyatlari_cek()
                if sayi > 0:
                    st.success(f"✅ {sayi} kripto fiyatı güncellendi!")
                    import time
                    time.sleep(1)
                    st.rerun()
                else:
                    st.warning("Kripto fiyatı çekilemedi (varlık tanımlı değil olabilir).")

        with oto_col4:
            if st.button("🥇 Altın Fiyatları", use_container_width=True):
                with st.spinner("Altın fiyatları hesaplanıyor..."):
                    sayi = altin_fiyatlari_cek()
                if sayi > 0:
                    st.success(f"✅ {sayi} altın fiyatı güncellendi!")
                    st.caption("ℹ️ Eritme değeri üzerinden hesaplandı (piyasa fiyatı %3-10 daha yüksek olabilir).")
                    import time
                    time.sleep(1)
                    st.rerun()
                else:
                    st.warning("Altın fiyatı çekilemedi.")

        with oto_col5:
            if st.button("💱 Döviz Kurları", use_container_width=True):
                bugun_str = date.today().strftime("%Y-%m-%d")
                with st.spinner("Yahoo Finance'tan USD/EUR/GBP kurları çekiliyor..."):
                    try:
                        kur_cek_ve_kaydet(bugun_str)
                        st.success("✅ Döviz kurları güncellendi!")
                    except Exception as e:
                        st.warning(f"Kur çekilemedi: {e}")
                import time
                time.sleep(1)
                st.rerun()

        # --- Satır 2: Tümünü Çek (kur dahil) ---
        if st.button("🔄 Tümünü Çek (Hisse + BIST + Kripto + Altın + Kur)", use_container_width=True, type="primary"):
            with st.spinner("Tüm fiyatlar ve kurlar çekiliyor..."):
                bugun_str = date.today().strftime("%Y-%m-%d")
                try:
                    kur_cek_ve_kaydet(bugun_str)
                except Exception as e:
                    st.warning(f"Kur güncellenemedi: {e}")
                sayi = tum_fiyatlari_cek()
            if sayi > 0:
                st.success(f"✅ Toplam {sayi} fiyat güncellendi, kurlar da yenilendi!")
                import time
                time.sleep(1)
                st.rerun()
            else:
                st.warning("Hiçbir fiyat çekilemedi.")

        # ==========================================
        # GEÇMİŞ VERİ TAMAMLA
        # ==========================================
        st.markdown("---")
        st.subheader("📅 Geçmiş Veri Tamamla")
        st.caption("Belirtilen tarihten bugüne günlük fiyat geçmişini Yahoo Finance'tan çeker. "
                   "Yeni varlık ekledikten sonra sadece ilgili türü çekmeniz yeterlidir.")

        from datetime import date as _date
        from dateutil.relativedelta import relativedelta as _rd

        gecmis_baslangic = st.date_input(
            "Başlangıç tarihi",
            value=_date.today() - _rd(years=1),
            key="gecmis_baslangic"
        )
        baslangic_str = str(gecmis_baslangic)

        # --- Satır 1: Yabancı Hisse, BIST, Kripto, Altın, Döviz Kuru ---
        gc1, gc2, gc3, gc4, gc5 = st.columns(5)

        with gc1:
            if st.button("📈 Yabancı Hisse Geçmiş", use_container_width=True, key="gecmis_hisse"):
                with st.spinner(f"Yabancı hisse geçmişi çekiliyor ({baslangic_str} →)..."):
                    sayi = hisse_fiyatlari_cek(baslangic_str)
                if sayi > 0:
                    st.success(f"✅ {sayi} kayıt eklendi!")
                    import time; time.sleep(1); st.rerun()
                else:
                    st.warning("Veri çekilemedi.")

        with gc2:
            if st.button("🏦 BIST Geçmiş", use_container_width=True, key="gecmis_bist"):
                with st.spinner(f"BIST geçmişi çekiliyor ({baslangic_str} →)..."):
                    sayi = bist_fiyatlari_cek(baslangic_str)
                if sayi > 0:
                    st.success(f"✅ {sayi} kayıt eklendi!")
                    import time; time.sleep(1); st.rerun()
                else:
                    st.warning("Veri çekilemedi.")

        with gc3:
            if st.button("🪙 Kripto Geçmiş", use_container_width=True, key="gecmis_kripto"):
                with st.spinner(f"Kripto geçmişi çekiliyor ({baslangic_str} →)..."):
                    sayi = kripto_fiyatlari_cek(baslangic_str)
                if sayi > 0:
                    st.success(f"✅ {sayi} kayıt eklendi!")
                    import time; time.sleep(1); st.rerun()
                else:
                    st.warning("Veri çekilemedi.")

        with gc4:
            if st.button("🥇 Altın Geçmiş", use_container_width=True, key="gecmis_altin"):
                with st.spinner(f"Altın geçmişi hesaplanıyor ({baslangic_str} →)..."):
                    sayi = altin_fiyatlari_cek(baslangic_str)
                if sayi > 0:
                    st.success(f"✅ {sayi} kayıt eklendi!")
                    import time; time.sleep(1); st.rerun()
                else:
                    st.warning("Veri çekilemedi.")

        with gc5:
            if st.button("💱 Kur Geçmişi", use_container_width=True, key="gecmis_kur"):
                with st.spinner(f"USD/EUR/GBP kur geçmişi çekiliyor ({baslangic_str} →)..."):
                    try:
                        kur_cek_ve_kaydet(baslangic_str)
                        st.success("✅ Kur geçmişi eklendi!")
                    except Exception as e:
                        st.warning(f"Kur çekilemedi: {e}")
                import time; time.sleep(1); st.rerun()

        # --- Satır 2: Tümü (kur dahil) ---
        if st.button("🔄 Tüm Geçmiş Verileri Çek (Fiyat + Kur)", use_container_width=True, key="gecmis_tum"):
            with st.spinner(f"{baslangic_str} tarihinden bugüne TÜM fiyatlar ve kurlar çekiliyor... (bu birkaç dakika sürebilir)"):
                try:
                    kur_cek_ve_kaydet(baslangic_str)
                except Exception as e:
                    st.warning(f"Kur güncellenemedi: {e}")
                sayi = tum_fiyatlari_cek(baslangic_str)
            if sayi > 0:
                st.success(f"✅ Toplam {sayi} geçmiş fiyat kaydı eklendi, kurlar da güncellendi!")
                import time; time.sleep(1); st.rerun()
            else:
                st.warning("Geçmiş veri çekilemedi.")

        # ==========================================
        # DOSYADAN YÜKLE (TEFAS CSV)
        # ==========================================
        st.markdown("---")
        st.subheader("📂 Dosyadan Yükle")
        st.caption("TEFAS'tan indirdiğiniz aylık CSV dosyalarını `data/tefas/` klasörüne koyun, ardından butona basın.")

        if st.button("📊 TEFAS Fon Fiyatları (CSV'den İçe Aktar)", use_container_width=True):
            with st.spinner("data/tefas/ klasöründeki CSV dosyaları okunuyor..."):
                try:
                    tefas_import()
                    st.success("✅ TEFAS fon fiyatları güncellendi!")
                    import time
                    time.sleep(1)
                    st.rerun()
                except Exception as e:
                    st.error(f"TEFAS import hatası: {e}")

        st.markdown("---")
        st.subheader("✏️ Manuel Fiyat Güncelle")
        st.info("💡 Fiyatları güncelleyip 'Fiyatları Kaydet' butonuna basın.")

        guncel_fiyatlar = {}

        # Mevduat türlerini filtrele (fiyatları her zaman 1, manuel güncellemeye gerek yok)
        manuel_df = df[~df["tur"].isin(MEVDUAT_TURLERI)].copy()

        if manuel_df.empty:
            st.info("Manuel fiyat girilecek varlık yok.")
        else:
            # Varlık türüne göre grupla
            for tur in sorted(manuel_df["tur"].unique()):
                tur_df = manuel_df[manuel_df["tur"] == tur].reset_index(drop=True)
                st.subheader(f"📂 {tur}")

                # İki kolon
                col1, col2 = st.columns(2)
                for i, (_, row) in enumerate(tur_df.iterrows()):
                    onceki = son_fiyatlar[son_fiyatlar["varlik_id"] == row["id"]]
                    onceki_fiyat = float(onceki["fiyat"].values[0]) if not onceki.empty else 0.0
                    onceki_tarih = onceki["tarih"].values[0] if not onceki.empty else "—"

                    hedef_col = col1 if i % 2 == 0 else col2
                    with hedef_col:
                        fiyat = st.number_input(
                            f"{row['kod']} ({onceki_tarih})",
                            min_value=0.0,
                            step=0.0001,
                            value=onceki_fiyat,
                            format="%.4f",
                            key=f"fiyat_{row['kod']}"
                        )
                        guncel_fiyatlar[row["kod"]] = {"fiyat": fiyat, "varlik_id": int(row["id"])}

                st.markdown("---")

        if st.button("💾 Fiyatları Kaydet", type="primary"):
            st.warning("⚠️ Aynı tarihte otomatik çekilmiş fiyat varsa üzerine yazılır.")
            baglanti = veritabani_baglan()
            cursor = baglanti.cursor()
            bugun = date.today().strftime("%Y-%m-%d")

            for kod, bilgi in guncel_fiyatlar.items():
                cursor.execute("""
                    INSERT OR REPLACE INTO fiyat_gecmisi (varlik_id, tarih, fiyat, kaynak)
                    VALUES (?, ?, ?, ?)
                """, (bilgi["varlik_id"], bugun, bilgi["fiyat"], "manuel"))

            baglanti.commit()
            senkronize_et()
            st.success("✅ Fiyatlar kaydedildi!")
            import time
            time.sleep(1)
            st.rerun()
